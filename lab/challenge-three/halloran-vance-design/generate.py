#!/usr/bin/env python3
"""
generate.py -- Halloran & Vance Design Partners, challenge-three Stage B.

Deterministic single source of truth: builds ledger.jsonl, documents.jsonl,
statements.jsonl, opening_position.json under lab/challenge-three/
halloran-vance-design/, renders the shipped corpus under
content/21-challenges/materials/challenge-three/halloran-vance-design/,
and writes answer-key.md.

Run: python3 lab/challenge-three/halloran-vance-design/generate.py
"""

from __future__ import annotations

import os
import random
import re
import shutil
import sys
import tempfile

HERE = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.abspath(os.path.join(HERE, "..", "..", ".."))
LAB_DIR = HERE
SLUG = "halloran-vance-design"
MATERIALS_ROOT = os.path.join(
    REPO_ROOT, "content", "21-challenges", "materials", "challenge-three", SLUG
)
MATERIALS_REL = f"content/21-challenges/materials/challenge-three/{SLUG}"

sys.path.insert(0, os.path.join(REPO_ROOT, "lab", "challenge-three"))
from lib import ledger as L  # noqa: E402
from lib import render as R  # noqa: E402

from PIL import Image, ImageDraw, ImageFont  # noqa: E402

SEED = 340716
RNG = random.Random(SEED)

# ---------------------------------------------------------------------------
# Clean slate (idempotent regeneration) -- only our own two trees.
# ---------------------------------------------------------------------------
if os.path.isdir(MATERIALS_ROOT):
    shutil.rmtree(MATERIALS_ROOT)
os.makedirs(MATERIALS_ROOT, exist_ok=True)
for fn in ("ledger.jsonl", "documents.jsonl", "statements.jsonl", "opening_position.json"):
    p = os.path.join(LAB_DIR, fn)
    if os.path.isfile(p):
        os.remove(p)

TMP_DIR = tempfile.mkdtemp(prefix="hv-gen-")

# ---------------------------------------------------------------------------
# Static entity data (all invented)
# ---------------------------------------------------------------------------

FIRM_NAME = "Halloran & Vance Design Partners"
FIRM_ADDR = "155 West 23rd Street, Suite 8B, New York, NY 10011"
FIRM_PHONE = "(212) 555-0148"
FIRM_EIN = "99-4471286"
FIRM_DOMAIN = "hvdesignpartners.com"

HALLORAN_NAME = "Margaret Halloran"
VANCE_NAME = "Owen Vance"
HALLORAN_EMAIL = "mhalloran@hvdesignpartners.com"
VANCE_EMAIL = "ovance@hvdesignpartners.com"
EMPLOYEE_NAME = "Priya Nair"
EMPLOYEE_TITLE = "Senior Designer"

BANK_NAME = "Hudson Yards Trust Company"
BANK_ADDR = "350 Tenth Avenue, New York, NY 10001"
BANK_ACCT_MASK = "****2289"
BANK_PHONE = "(212) 555-0177"

PRIOR_CPA = "Ostrander & Cole CPAs LLP"
PRIOR_CPA_ADDR = "245 Fifth Avenue, Suite 1100, New York, NY 10016"
PRIOR_CPA_PHONE = "(212) 555-0173"
PRIOR_CPA_EMAIL = "closing@ostrandercolecpas.com"

BOOKKEEPER = "Fillmore Bookkeeping & Tax LLC"
BOOKKEEPER_EMAIL = "fillmore@fillmorebooks.com"
BOOKKEEPER_ADDR = "88 Pine Street, New York, NY 10005"

PAYROLL_PROVIDER = "Empire State Payroll Partners"
PAYROLL_PROVIDER_ADDR = "90 Broad Street, New York, NY 10004"
PAYROLL_PROVIDER_EMAIL = "notices@empirestatepayroll.com"

LANDLORD = "23rd Street Flatiron Holdings LLC"
UTILITY_CO = "Empire Gas & Electric Co."
TELECOM_CO = "Gotham Fiber & Voice"
INSURER = "Gramercy Professional Insurance Group"
OFFICE_SUPPLY_VENDOR = "Tribeca Paper & Print Co"
HARDWARE_VENDOR = "Chelsea Hardware & Supply"
AD_VENDOR = "Flatiron Design Quarterly"

REYES = "Reyes Drafting Studio"
IONESCU = "Ionescu Lighting Consultants"
COBALT_STRUCTURAL = "Cobalt Structural Engineering PLLC"
RENDERCRAFT = "Rendercraft Visualization Studio"
GRAMERCY_PRINT = "Gramercy Print & Copy"
SOFTWARE_VENDOR = "DraftLine Software Solutions"
STORAGE_VENDOR = "Chelsea Mini Storage"
COURIER_VENDOR = "Flatiron Rush Courier"
FRAME_VENDOR = "Bowery Frame & Mount"
DUES_ORG = "New York Society of Interior Design Professionals"
SECOND_INSURER = "Hudson Professional Liability Underwriters"

CLIENTS = {
    "bellcourt": "Bellcourt Retail Group",
    "ansel": "Ansel Family Residence",
    "larkspur": "Larkspur Hospitality LLC",
    "whitfield": "Whitfield Family Trust",
    "meridian": "Meridian Law Offices",
    "pemberton": "Pemberton Townhouse",
    "fenwick": "Fenwick Hospitality Group",
}

PERIOD_START = "2024-07-01"
PERIOD_END = "2025-06-30"
AS_OF_OPEN = "2024-06-30"
DOC_SPAN_END = "2025-08-31"

MONTHS = [
    "2024-07", "2024-08", "2024-09", "2024-10", "2024-11", "2024-12",
    "2025-01", "2025-02", "2025-03", "2025-04", "2025-05", "2025-06",
    "2025-07", "2025-08",
]

MONTH_LABEL = {
    "2024-07": "July 2024", "2024-08": "August 2024", "2024-09": "September 2024",
    "2024-10": "October 2024", "2024-11": "November 2024", "2024-12": "December 2024",
    "2025-01": "January 2025", "2025-02": "February 2025", "2025-03": "March 2025",
    "2025-04": "April 2025", "2025-05": "May 2025", "2025-06": "June 2025",
    "2025-07": "July 2025", "2025-08": "August 2025",
}


def month_bounds(m: str) -> tuple[str, str]:
    import calendar
    y, mo = int(m[:4]), int(m[5:7])
    last = calendar.monthrange(y, mo)[1]
    return f"{m}-01", f"{m}-{last:02d}"


def cents(dollars: float) -> int:
    return round(dollars * 100)


# ---------------------------------------------------------------------------
# ID counters
# ---------------------------------------------------------------------------

_counters: dict[str, int] = {}


def next_id(prefix: str) -> str:
    _counters[prefix] = _counters.get(prefix, 0) + 1
    return f"{prefix}-{_counters[prefix]:03d}"


# ---------------------------------------------------------------------------
# Registries
# ---------------------------------------------------------------------------

LEDGER: list[dict] = []
# entry_id -> bank-style line description, for entries whose ledger memo
# carries internal bookkeeping reasoning that no bank would ever print on a
# statement. Ledger memos stay exactly as SPEC 1.1 wants them (plain-language
# explanation of the line); this dict only affects the *statement view*.
BANK_DESC: dict[str, str] = {}
DOCUMENTS: list[dict] = []


def add_line(entry_id: str, date_: str, code: str, debit: int, credit: int,
             memo: str, counterparty: str, doc_ids: list[str]) -> None:
    assert (debit == 0) != (credit == 0) or (debit == 0 and credit == 0 and False), \
        "exactly one of debit/credit must be non-zero"
    assert doc_ids, f"line on {code} ({entry_id}) has no doc_ids"
    LEDGER.append({
        "entry_id": entry_id,
        "date": date_,
        "account_code": code,
        "account_name": L.CHART[code]["name"],
        "debit": debit,
        "credit": credit,
        "memo": memo,
        "counterparty": counterparty,
        "doc_ids": list(doc_ids),
    })


def add_entry(entry_id: str, date_: str, lines: list[tuple], ) -> None:
    """lines: list of (code, debit, credit, memo, counterparty, doc_ids)"""
    td = sum(l[1] for l in lines)
    tc = sum(l[2] for l in lines)
    assert td == tc, f"entry {entry_id} unbalanced: debit {td} != credit {tc}"
    for code, debit, credit, memo, counterparty, doc_ids in lines:
        add_line(entry_id, date_, code, debit, credit, memo, counterparty, doc_ids)


def add_doc(doc_id: str, kind: str, path_rel: str, fmt: str, scanned: bool,
            issued_date: str, counterparty: str, amount: int | None) -> None:
    DOCUMENTS.append({
        "doc_id": doc_id,
        "kind": kind,
        "path": path_rel,
        "format": fmt,
        "scanned": scanned,
        "issued_date": issued_date,
        "counterparty": counterparty,
        "amount": amount,
    })


# ---------------------------------------------------------------------------
# Batch (email) schedule
# ---------------------------------------------------------------------------

BATCH_SPECS = [
    ("opening", "2024-07-03", "closing-financials-june-30"),
    ("monthly-2024-07", "2024-07-29", "invoices-payroll-july"),
    ("monthly-2024-08", "2024-08-28", "invoices-bills-payroll-august"),
    ("stmt-1", "2024-09-05", "bank-statements-july-august"),
    ("monthly-2024-09", "2024-09-27", "invoices-bills-payroll-september"),
    ("expense-log-4", "2024-08-14", "july-expense-log"),
    ("expense-log-1", "2024-10-08", "september-expense-log"),
    ("monthly-2024-10", "2024-10-30", "invoices-bills-october"),
    ("stmt-2", "2024-11-08", "bank-statements-sept-oct"),
    ("monthly-2024-11", "2024-11-26", "invoices-bills-payroll-november"),
    ("expense-log-5", "2024-12-18", "november-expense-log"),
    ("monthly-2024-12", "2024-12-23", "whitfield-milestone-invoice-capital-december"),
    ("stmt-3", "2025-01-10", "bank-statements-nov-dec"),
    ("monthly-2025-01", "2025-01-29", "invoices-bills-payroll-january"),
    ("expense-log-2", "2025-02-06", "january-expense-log"),
    ("monthly-2025-02", "2025-02-26", "invoices-bills-payroll-february"),
    ("stmt-4", "2025-03-12", "bank-statements-jan-feb"),
    ("expense-log-6", "2025-03-14", "february-expense-log"),
    ("monthly-2025-03", "2025-03-27", "invoices-bills-payroll-march"),
    ("monthly-2025-04", "2025-04-29", "invoices-bills-payroll-april"),
    ("expense-log-3", "2025-05-09", "april-expense-log"),
    ("stmt-5", "2025-05-14", "bank-statements-mar-apr"),
    ("monthly-2025-05", "2025-05-28", "invoices-bills-payroll-may"),
    ("monthly-2025-06", "2025-06-26", "invoices-bills-payroll-june"),
    ("stmt-6", "2025-07-09", "bank-statements-may-june"),
    ("expense-log-7", "2025-07-16", "june-expense-log"),
    ("monthly-2025-07", "2025-07-30", "invoices-payroll-july-stub"),
    ("monthly-2025-08", "2025-08-27", "invoices-payroll-august-stub"),
    ("stmt-7", "2025-09-08", "bank-statements-july-august-stub"),
]

_sorted_specs = sorted(BATCH_SPECS, key=lambda t: (t[1], t[0]))
BATCH_DIR: dict[str, str] = {}
BATCH_DATE: dict[str, str] = {}
for i, (key, date_, slug) in enumerate(_sorted_specs, start=1):
    dirname = f"{i:06d}_{date_}_{slug}"
    BATCH_DIR[key] = dirname
    BATCH_DATE[key] = date_

BATCH_EMAIL: dict[str, dict] = {key: {"from": "", "to": "", "subject": "", "paragraphs": [], "attachments": []}
                                 for key in BATCH_DIR}


def batch_path(key: str, filename: str) -> str:
    d = os.path.join(MATERIALS_ROOT, BATCH_DIR[key])
    os.makedirs(d, exist_ok=True)
    return os.path.join(d, filename)


def batch_relpath(key: str, filename: str) -> str:
    return f"{MATERIALS_REL}/{BATCH_DIR[key]}/{filename}"


def batch_attach(key: str, filename: str, note: str) -> None:
    BATCH_EMAIL[key]["attachments"].append((filename, note))


def batch_para(key: str, text: str) -> None:
    BATCH_EMAIL[key]["paragraphs"].append(text)


# ---------------------------------------------------------------------------
# HTML / rendering helpers
# ---------------------------------------------------------------------------

BASE_CSS = """
@page { size: letter; margin: 0.7in; }
body { font-family: Helvetica, Arial, sans-serif; font-size: 10.5pt; color: #1a1a1a; }
h1 { font-size: 15pt; margin-bottom: 2px; }
h2 { font-size: 12pt; margin-top: 18px; }
.header { margin-bottom: 14px; }
.small { font-size: 9pt; color: #444; }
table { border-collapse: collapse; width: 100%; margin-top: 8px; }
th, td { border: 1px solid #999; padding: 4px 6px; font-size: 9.5pt; text-align: left; }
th { background: #eee; }
.right { text-align: right; }
.total-row td { font-weight: bold; }
"""


def html_wrap(body: str) -> str:
    return f"<html><head><meta charset='utf-8'></head><body>{body}</body></html>"


def money_str(c: int) -> str:
    sign = "-" if c < 0 else ""
    c = abs(c)
    return f"{sign}${c // 100:,}.{c % 100:02d}"


def render_pdf(key: str, filename: str, html_body: str, kind: str, issued_date: str,
               counterparty: str, amount: int | None, doc_id: str | None = None,
               fmt_note: str = "") -> str:
    doc_id = doc_id or next_id("DOC")
    out_path = batch_path(key, filename)
    R.render_html_to_pdf(html_wrap(html_body), out_path, css=BASE_CSS)
    rel = batch_relpath(key, filename)
    add_doc(doc_id, kind, rel, "pdf", False, issued_date, counterparty, amount)
    return doc_id


def render_pdf_scanned(key: str, filename: str, html_body: str, kind: str, issued_date: str,
                        counterparty: str, amount: int | None, seed: int,
                        doc_id: str | None = None) -> str:
    """Render to a text PDF in TMP_DIR, then scanify() it into an image-only
    PDF at the real shipped path -- kept out of the materials tree so the
    text-layer source never leaks next to a document declared scanned=True."""
    doc_id = doc_id or next_id("DOC")
    tmp_pdf = os.path.join(TMP_DIR, f"{doc_id}_src.pdf")
    R.render_html_to_pdf(html_wrap(html_body), tmp_pdf, css=BASE_CSS)
    out_path = batch_path(key, filename)
    R.scanify(tmp_pdf, out_path, seed=seed)
    rel = batch_relpath(key, filename)
    add_doc(doc_id, kind, rel, "pdf", True, issued_date, counterparty, amount)
    return doc_id


def render_docx_doc(key: str, filename: str, title: str, paragraphs: list[str],
                     table_rows: list[list[str]] | None, kind: str, issued_date: str,
                     counterparty: str, amount: int | None, doc_id: str | None = None) -> str:
    doc_id = doc_id or next_id("DOC")
    out_path = batch_path(key, filename)
    R.render_docx(out_path, title, paragraphs, table_rows=table_rows)
    rel = batch_relpath(key, filename)
    add_doc(doc_id, kind, rel, "docx", False, issued_date, counterparty, amount)
    return doc_id


def render_xlsx_doc(key: str, filename: str, sheets: dict, kind: str, issued_date: str,
                     counterparty: str, amount: int | None, doc_id: str | None = None) -> str:
    doc_id = doc_id or next_id("DOC")
    out_path = batch_path(key, filename)
    R.render_xlsx(out_path, sheets)
    rel = batch_relpath(key, filename)
    add_doc(doc_id, kind, rel, "xlsx", False, issued_date, counterparty, amount)
    return doc_id


def render_txt_doc(key: str, filename: str, text: str, kind: str, issued_date: str,
                    counterparty: str, amount: int | None, doc_id: str | None = None) -> str:
    doc_id = doc_id or next_id("DOC")
    out_path = batch_path(key, filename)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(text)
    rel = batch_relpath(key, filename)
    add_doc(doc_id, kind, rel, "txt", False, issued_date, counterparty, amount)
    return doc_id


def make_flat_receipt_png(lines: list[str], path: str, width: int = 1100, height: int = 820) -> None:
    img = Image.new("RGB", (width, height), (252, 251, 245))
    d = ImageDraw.Draw(img)
    try:
        font = ImageFont.truetype("DejaVuSans.ttf", 34)
    except OSError:
        font = ImageFont.load_default()
    y = 40
    for line in lines:
        d.text((45, y), line, font=font, fill=(10, 10, 10))
        y += 52
    img.save(path)


def render_receipt_jpg(key: str, filename: str, lines: list[str], kind: str, issued_date: str,
                        counterparty: str, amount: int | None, seed: int,
                        doc_id: str | None = None) -> str:
    doc_id = doc_id or next_id("DOC")
    flat_path = os.path.join(TMP_DIR, f"{doc_id}_flat.png")
    make_flat_receipt_png(lines, flat_path)
    out_path = batch_path(key, filename)
    R.photograph_receipt(flat_path, out_path, seed=seed)
    rel = batch_relpath(key, filename)
    add_doc(doc_id, kind, rel, "jpg", True, issued_date, counterparty, amount)
    return doc_id


def render_handwritten_jpg(key: str, filename: str, lines: list[str], kind: str, issued_date: str,
                            counterparty: str, amount: int | None, seed: int,
                            doc_id: str | None = None) -> str:
    doc_id = doc_id or next_id("DOC")
    out_path = batch_path(key, filename)
    R.handwritten_note_image(lines, out_path, seed=seed)
    rel = batch_relpath(key, filename)
    add_doc(doc_id, kind, rel, "jpg", True, issued_date, counterparty, amount)
    return doc_id


# ---------------------------------------------------------------------------
# Statement doc id pre-allocation (path/kind known now; content rendered later)
# ---------------------------------------------------------------------------

STMT_BATCH_FOR_MONTH = {
    "2024-07": "stmt-1", "2024-08": "stmt-1",
    "2024-09": "stmt-2", "2024-10": "stmt-2",
    "2024-11": "stmt-3", "2024-12": "stmt-3",
    "2025-01": "stmt-4", "2025-02": "stmt-4",
    "2025-03": "stmt-5", "2025-04": "stmt-5",
    "2025-05": "stmt-6", "2025-06": "stmt-6",
    "2025-07": "stmt-7", "2025-08": "stmt-7",
}

STMT_DOC_ID = {m: f"DOC-STMT-{m}" for m in MONTHS}
STMT_FILENAME = {m: f"hudson_yards_trust_statement_{m}.pdf" for m in MONTHS}


def stmt_doc_id(iso_date: str) -> str:
    return STMT_DOC_ID[iso_date[:7]]


# ---------------------------------------------------------------------------
# Opening letter (prior CPA firm) + OB-1 entry + opening_position.json
# ---------------------------------------------------------------------------

OPEN_CASH = cents(48000)
OPEN_AR = {
    "Bellcourt Retail Group": cents(34000),
    "Ansel Family Residence": cents(21400),
    "Larkspur Hospitality LLC": cents(25000),
}
OPEN_AP = {
    "Reyes Drafting Studio": cents(6200),
    "Ionescu Lighting Consultants": cents(7200),
    "Tribeca Paper & Print Co": cents(3000),
}
OPEN_CAP_HALLORAN = cents(70000)
OPEN_CAP_VANCE = cents(42000)

assert OPEN_CASH + sum(OPEN_AR.values()) == sum(OPEN_AP.values()) + OPEN_CAP_HALLORAN + OPEN_CAP_VANCE

DOC_OPEN = "DOC-OPEN"

opening_letter_body = f"""
<div class="header">
<h1>{PRIOR_CPA}</h1>
<div class="small">{PRIOR_CPA_ADDR} &middot; {PRIOR_CPA_PHONE}</div>
</div>
<p>30 June 2024</p>
<p>{HALLORAN_NAME} and {VANCE_NAME}<br/>{FIRM_NAME}<br/>{FIRM_ADDR}</p>
<p>Dear Margaret and Owen,</p>
<p>As requested, this letter sets out the closing position of {FIRM_NAME} (the
"Partnership") as at 30 June 2024, the end of our engagement preparing your
books through the current fiscal year-end. We have not audited these
figures; they reflect the Partnership's own records as reconciled by this
firm.</p>
<p><strong>Cash.</strong> The Partnership held {money_str(OPEN_CASH)} in its
operating account at {BANK_NAME} (account ending {BANK_ACCT_MASK}) at the
close of business on 30 June 2024.</p>
<p><strong>Accounts receivable.</strong> Three clients owed a combined
{money_str(sum(OPEN_AR.values()))} for design fees billed but not yet
collected: {money_str(OPEN_AR["Bellcourt Retail Group"])} from Bellcourt
Retail Group, {money_str(OPEN_AR["Ansel Family Residence"])} from Ansel
Family Residence, and {money_str(OPEN_AR["Larkspur Hospitality LLC"])} from
Larkspur Hospitality LLC.</p>
<p><strong>Accounts payable.</strong> The Partnership owed a combined
{money_str(sum(OPEN_AP.values()))} to three vendors:
{money_str(OPEN_AP["Reyes Drafting Studio"])} to Reyes Drafting Studio for
drafting services, {money_str(OPEN_AP["Ionescu Lighting Consultants"])} to
Ionescu Lighting Consultants for lighting design services, and
{money_str(OPEN_AP["Tribeca Paper & Print Co"])} to Tribeca Paper &amp;
Print Co for office supplies.</p>
<p><strong>Partner capital.</strong> Individual capital accounts stood at
{money_str(OPEN_CAP_HALLORAN)} for Ms. Halloran and
{money_str(OPEN_CAP_VANCE)} for Mr. Vance, reflecting each partner's
cumulative contributions and prior earnings net of draws. Under the
partnership agreement, profits and losses are shared 60% to Ms. Halloran
and 40% to Mr. Vance; this ratio governs the allocation of income between
the partners and is unrelated to the capital account balances above.</p>
<p>The table below summarizes the position as at 30 June 2024:</p>
<table>
<tr><th>Item</th><th class="right">Amount</th></tr>
<tr><td>Cash - operating (Hudson Yards Trust Company, {BANK_ACCT_MASK})</td><td class="right">{money_str(OPEN_CASH)}</td></tr>
<tr><td>Accounts receivable (3 clients, named above)</td><td class="right">{money_str(sum(OPEN_AR.values()))}</td></tr>
<tr><td>Accounts payable (3 vendors, named above)</td><td class="right">({money_str(sum(OPEN_AP.values()))})</td></tr>
<tr><td>Partner capital - M. Halloran</td><td class="right">{money_str(OPEN_CAP_HALLORAN)}</td></tr>
<tr><td>Partner capital - O. Vance</td><td class="right">{money_str(OPEN_CAP_VANCE)}</td></tr>
</table>
<p>Please let us know if you have any questions about the figures above.
It has been a pleasure serving the Partnership these past years.</p>
<p>Sincerely,<br/>Diane Ostrander, CPA<br/>{PRIOR_CPA}</p>
"""

render_pdf("opening", "opening_letter.pdf", opening_letter_body, "opening_letter",
           "2024-06-30", PRIOR_CPA, None, doc_id=DOC_OPEN)
batch_attach("opening", "opening_letter.pdf", "closing position letter as at 30 June 2024")

# OB-1: opening balances, dated the day before period_start
ob_lines = [
    ("1000", OPEN_CASH, 0, "Opening cash - operating (Hudson Yards Trust Co.)", BANK_NAME, [DOC_OPEN]),
]
for debtor, amt in OPEN_AR.items():
    ob_lines.append(("1200", amt, 0, f"Opening AR - {debtor}", debtor, [DOC_OPEN]))
for creditor, amt in OPEN_AP.items():
    ob_lines.append(("2000", 0, amt, f"Opening AP - {creditor}", creditor, [DOC_OPEN]))
ob_lines.append(("3100", 0, OPEN_CAP_HALLORAN, "Opening partner capital - Halloran", HALLORAN_NAME, [DOC_OPEN]))
ob_lines.append(("3110", 0, OPEN_CAP_VANCE, "Opening partner capital - Vance", VANCE_NAME, [DOC_OPEN]))

add_entry("OB-1", "2024-06-30", ob_lines)

batch_para("opening",
    f"Attached is our closing letter for the fiscal year ended 30 June 2024, "
    f"including the balances we'll be handing off to whoever picks up the "
    f"bookkeeping going forward. Let us know if anything needs clarifying.")
BATCH_EMAIL["opening"]["from"] = f"{PRIOR_CPA} <{PRIOR_CPA_EMAIL}>"
BATCH_EMAIL["opening"]["to"] = f"{HALLORAN_NAME}, {VANCE_NAME} <{HALLORAN_EMAIL}>"
BATCH_EMAIL["opening"]["subject"] = "Closing financials as at June 30, 2024"

print("Opening position built.", file=sys.stderr)

# ---------------------------------------------------------------------------
# Invoices out (revenue recognized on issuance)
# ---------------------------------------------------------------------------

# (month_key, client_key, phase, amount_dollars, issue_day, paid: (date) or None, docx: bool)
INVOICES = [
    ("2024-07", "ansel", "Design Development - Phase 2", 10000, 8, "2024-08-19", False),
    ("2024-07", "pemberton", "Schematic Design", 8000, 22, "2024-08-27", True),

    ("2024-08", "meridian", "Construction Documents - Phase 1", 9000, 7, "2024-09-16", False),
    ("2024-08", "fenwick", "Concept Design - Location 1", 6000, 21, "2024-09-30", True),

    ("2024-09", "bellcourt", "Construction Documents", 26000, 6, "2024-10-18", False),
    ("2024-09", "ansel", "Construction Documents - Phase 1", 20000, 20, "2024-11-01", False),

    ("2024-10", "whitfield", "Design Development", 30000, 9, "2024-11-14", True),
    ("2024-10", "larkspur", "Schematic Design", 18000, 23, "2024-12-05", False),

    ("2024-11", "meridian", "Construction Documents - Phase 2", 24000, 8, "2024-12-19", False),
    ("2024-11", "pemberton", "Design Development", 18000, 22, "2025-01-06", True),

    ("2024-12", "whitfield", "Milestone: Substantial Completion - Phase 3", 185000, 12, "2025-01-24", False),
    ("2024-12", "fenwick", "Concept Design - Location 2", 25000, 20, "2025-01-31", False),

    ("2025-01", "bellcourt", "Construction Administration - Phase 1", 32000, 10, "2025-02-21", False),
    ("2025-01", "larkspur", "Design Development", 26000, 24, "2025-03-07", True),

    ("2025-02", "ansel", "Construction Administration", 28000, 7, "2025-03-19", False),
    ("2025-02", "meridian", "Construction Administration", 24000, 21, "2025-04-02", False),

    ("2025-03", "pemberton", "Construction Documents", 34000, 6, "2025-04-17", False),
    ("2025-03", "fenwick", "Construction Documents - Location 1", 26000, 20, "2025-05-01", True),

    ("2025-04", "bellcourt", "Construction Administration - Phase 2", 30000, 9, "2025-05-20", False),
    ("2025-04", "larkspur", "Construction Documents", 25000, 23, None, False),  # unpaid at period end

    ("2025-05", "whitfield", "Final Punch List & Closeout", 32000, 8, None, False),  # unpaid at period end
    ("2025-05", "pemberton", "Construction Administration", 28000, 22, None, False),  # unpaid at period end

    ("2025-06", "meridian", "Final Completion", 27000, 6, None, False),  # unpaid at period end
    ("2025-06", "fenwick", "Construction Administration - Location 2", 23000, 20, None, False),  # unpaid at period end

    ("2025-07", "ansel", "Punch List", 18000, 10, None, False),
    ("2025-08", "meridian", "Warranty Review", 15000, 8, None, False),
]

INV_NUM_START = 1041


def invoice_html(inv_num: str, client: str, phase: str, amount: int, issue_iso: str,
                  issue_display: str) -> str:
    due = issue_iso  # net 30 shown in terms line, not computed here
    return f"""
<div class="header">
<h1>{FIRM_NAME}</h1>
<div class="small">{FIRM_ADDR} &middot; {FIRM_PHONE} &middot; EIN {FIRM_EIN}</div>
</div>
<h2>Invoice {inv_num}</h2>
<p>Bill to: <strong>{client}</strong><br/>Date: {issue_display}<br/>Terms: Net 30</p>
<table>
<tr><th>Project Phase</th><th class="right">Amount</th></tr>
<tr><td>{phase}</td><td class="right">{money_str(amount)}</td></tr>
<tr class="total-row"><td>Total Due</td><td class="right">{money_str(amount)}</td></tr>
</table>
<p class="small">Please remit payment by check or wire to {FIRM_NAME},
referencing invoice {inv_num}. Thank you for the opportunity to work on
this project.</p>
"""


def fmt_date_variant(iso: str, variant: int) -> str:
    y, m, d = iso.split("-")
    months_full = ["January", "February", "March", "April", "May", "June", "July",
                   "August", "September", "October", "November", "December"]
    if variant == 0:
        return f"{m}/{d}/{y}"
    if variant == 1:
        return iso
    return f"{int(d)} {months_full[int(m) - 1][:3]} {y}"


for idx, (mkey, ckey, phase, amt_dollars, day, paid_date, as_docx) in enumerate(INVOICES):
    client = CLIENTS[ckey]
    amount = cents(amt_dollars)
    issue_iso = f"{mkey}-{day:02d}"
    inv_num = f"HV-{INV_NUM_START + idx}"
    variant = idx % 3
    issue_display = fmt_date_variant(issue_iso, variant)
    doc_id = next_id("DOC-INV")
    batch_key = f"monthly-{mkey}"
    fname_base = f"invoice_{inv_num.lower().replace('-', '')}_{ckey}"
    if as_docx:
        paras = [
            f"{FIRM_NAME}",
            f"{FIRM_ADDR}  |  {FIRM_PHONE}  |  EIN {FIRM_EIN}",
            "",
            f"Invoice {inv_num}",
            f"Bill to: {client}",
            f"Date: {issue_display}",
            "Terms: Net 30",
        ]
        rows = [["Project Phase", "Amount"], [phase, money_str(amount)], ["Total Due", money_str(amount)]]
        render_docx_doc(batch_key, f"{fname_base}.docx", f"Invoice {inv_num}", paras, rows,
                         "invoice_out", issue_display, client, amount, doc_id=doc_id)
        fname = f"{fname_base}.docx"
    else:
        render_pdf(batch_key, f"{fname_base}.pdf", invoice_html(inv_num, client, phase, amount, issue_iso, issue_display),
                    "invoice_out", issue_display, client, amount, doc_id=doc_id)
        fname = f"{fname_base}.pdf"
    batch_attach(batch_key, fname, f"invoice {inv_num} to {client} - {phase} - {money_str(amount)}")

    entry_id = next_id("INV")
    add_entry(entry_id, issue_iso, [
        ("1200", amount, 0, f"Invoice {inv_num} - {client} - {phase}", client, [doc_id]),
        ("4000", 0, amount, f"Invoice {inv_num} - {client} - {phase}", client, [doc_id]),
    ])

    if paid_date:
        pmt_entry = next_id("PMT")
        sd = stmt_doc_id(paid_date)
        add_entry(pmt_entry, paid_date, [
            ("1000", amount, 0, f"Payment received - {client} - Invoice {inv_num}", client, [sd]),
            ("1200", 0, amount, f"Payment received - {client} - Invoice {inv_num}", client, [doc_id]),
        ])

print(f"Invoices built: {len(INVOICES)}", file=sys.stderr)

# A handful of extra, specific sentences so the monthly batches don't all
# read as the same canned template -- real detail pulled from the actual
# billing/vendor data already defined above.
batch_para("monthly-2024-07",
    "Quiet one this month, as usual for July -- just the Ansel and Pemberton phases going out.")
batch_para("monthly-2024-08",
    "August's still slow. Things should pick back up once Bellcourt's CDs are ready in September.")
batch_para("monthly-2024-12",
    "Big one this month -- Whitfield signed off on substantial completion, so the "
    "Phase 3 milestone invoice finally went out. That one's been a long time coming.")
batch_para("monthly-2025-01",
    "Heads up, Reyes's bill this month is a bit higher than usual ($3,400 vs. the "
    "usual ~$2,800) -- they picked up extra drafting hours over the holidays for "
    "the Whitfield closeout set. Nothing to flag, just didn't want it to look odd.")
batch_para("monthly-2025-06",
    "Also -- Larkspur still hasn't paid the April invoice ($25,000). I'll follow up "
    "with their office next week before we chase it further.")

# ---------------------------------------------------------------------------
# Opening AR / AP settlement (in-period; excluded from current revenue/expense)
# ---------------------------------------------------------------------------

OPEN_AR_SETTLE_DATES = {
    "Bellcourt Retail Group": "2024-07-18",
    "Ansel Family Residence": "2024-08-09",
    "Larkspur Hospitality LLC": "2024-09-13",
}
OPEN_AP_SETTLE_DATES = {
    "Reyes Drafting Studio": "2024-07-15",
    "Ionescu Lighting Consultants": "2024-08-05",
    "Tribeca Paper & Print Co": "2024-07-22",
}

for debtor, amt in OPEN_AR.items():
    d = OPEN_AR_SETTLE_DATES[debtor]
    eid = next_id("OBSETTLE-AR")
    sd = stmt_doc_id(d)
    BANK_DESC[eid] = f"DEPOSIT - {debtor.upper()}"
    add_entry(eid, d, [
        ("1000", amt, 0, f"Collection of opening AR - {debtor} (per prior CPA closing letter)", debtor, [sd, DOC_OPEN]),
        ("1200", 0, amt, f"Collection of opening AR - {debtor} (per prior CPA closing letter)", debtor, [DOC_OPEN]),
    ])

for creditor, amt in OPEN_AP.items():
    d = OPEN_AP_SETTLE_DATES[creditor]
    eid = next_id("OBSETTLE-AP")
    sd = stmt_doc_id(d)
    BANK_DESC[eid] = f"CHECK PAID - {creditor.upper()}"
    add_entry(eid, d, [
        ("2000", amt, 0, f"Payment of opening AP - {creditor} (per prior CPA closing letter)", creditor, [DOC_OPEN]),
        ("1000", 0, amt, f"Payment of opening AP - {creditor} (per prior CPA closing letter)", creditor, [sd, DOC_OPEN]),
    ])

print("Opening AR/AP settlements built.", file=sys.stderr)

# ---------------------------------------------------------------------------
# Subcontractor & vendor bills (accounts payable) -- bundled per month into
# a single PDF attachment where more than one bill lands that month, the way
# a real client actually scans/forwards a stack of paperwork together. This
# is what makes the corpus dense: five sub-consultants and vendors billing
# through the same handful of monthly attachments rather than one file per
# invoice, exactly the "document lives where you don't expect it" trap the
# corpus needs (a given vendor's bill is page N of that month's bundle, not
# its own file).
# ---------------------------------------------------------------------------

COBALT_ADDR = "145 Hudson Street, New York, NY 10013"
RENDERCRAFT_ADDR = "20 Jay Street, Brooklyn, NY 11201"
GRAMERCY_PRINT_ADDR = "301 Park Avenue South, New York, NY 10010"
SOFTWARE_VENDOR_ADDR = "PO Box 4471, New York, NY 10008"
CODE_CONSULTANT = "Third Rail Code & Egress Consulting"
CODE_CONSULTANT_ADDR = "88 Leonard Street, New York, NY 10013"
BELLWEATHER = "Bellweather Procurement & FF&E Sourcing"
BELLWEATHER_ADDR = "224 West 30th Street, New York, NY 10001"
EQUIPMENT_VENDOR = "Metro Office Equipment Co"
EQUIPMENT_VENDOR_ADDR = "39 West 19th Street, New York, NY 10011"

VENDOR_ADDR = {
    REYES: "412 Atlantic Avenue, Brooklyn, NY 11217",
    IONESCU: "77 Greene Street, New York, NY 10012",
    BOOKKEEPER: BOOKKEEPER_ADDR,
    COBALT_STRUCTURAL: COBALT_ADDR,
    RENDERCRAFT: RENDERCRAFT_ADDR,
    GRAMERCY_PRINT: GRAMERCY_PRINT_ADDR,
    SOFTWARE_VENDOR: SOFTWARE_VENDOR_ADDR,
    CODE_CONSULTANT: CODE_CONSULTANT_ADDR,
    AD_VENDOR: "Advertising Sales, 41 Union Square W, New York, NY 10003",
    BELLWEATHER: BELLWEATHER_ADDR,
    EQUIPMENT_VENDOR: EQUIPMENT_VENDOR_ADDR,
}
VENDOR_DESC = {
    REYES: "CAD drafting and construction-document support",
    IONESCU: "Lighting design and specification services",
    BOOKKEEPER: "Bookkeeping services",
    COBALT_STRUCTURAL: "Structural engineering consultation",
    RENDERCRAFT: "3D visualization and rendering services",
    GRAMERCY_PRINT: "Large-format reprographics and plotting",
    SOFTWARE_VENDOR: "Annual CAD/rendering software subscription plan",
    CODE_CONSULTANT: "Code and egress compliance review",
    AD_VENDOR: "Print advertising placement",
    BELLWEATHER: "FF&E procurement and vendor sourcing coordination",
    EQUIPMENT_VENDOR: "Large-format plotter purchase",
}

# (vendor, issue_date, amount_dollars, paid_date_or_None, account_code)
BILLS = [
    # Reyes Drafting Studio -- regular, scales up with drafting-heavy CD phases
    (REYES, "2024-07-20", 3200, "2024-08-15", "6040"),
    (REYES, "2024-08-19", 2800, "2024-09-15", "6040"),
    (REYES, "2024-09-18", 6200, "2024-10-15", "6040"),
    (REYES, "2024-10-21", 7400, "2024-11-18", "6040"),
    (REYES, "2024-11-19", 8600, "2024-12-16", "6040"),
    (REYES, "2024-12-16", 7200, "2025-01-14", "6040"),
    (REYES, "2025-01-17", 6800, "2025-02-13", "6040"),
    (REYES, "2025-02-18", 6200, "2025-03-14", "6040"),
    (REYES, "2025-03-19", 7800, "2025-04-16", "6040"),
    (REYES, "2025-04-18", 7000, "2025-05-15", "6040"),
    (REYES, "2025-05-20", 6600, None, "6040"),
    (REYES, "2025-06-18", 5800, None, "6040"),
    (REYES, "2025-07-21", 3400, "2025-08-18", "6040"),
    (REYES, "2025-08-19", 3000, None, "6040"),

    # Ionescu Lighting Consultants -- regular, skips the deepest summer lull
    (IONESCU, "2024-07-22", 2400, "2024-08-19", "6040"),
    (IONESCU, "2024-09-16", 3800, "2024-10-14", "6040"),
    (IONESCU, "2024-10-25", 5600, "2024-11-21", "6040"),
    (IONESCU, "2024-11-22", 6800, "2024-12-19", "6040"),
    (IONESCU, "2024-12-18", 7600, "2025-01-16", "6040"),
    (IONESCU, "2025-01-24", 5200, "2025-02-20", "6040"),
    (IONESCU, "2025-02-19", 4600, "2025-03-18", "6040"),
    (IONESCU, "2025-03-21", 5800, "2025-04-17", "6040"),
    (IONESCU, "2025-04-22", 5200, "2025-05-19", "6040"),
    (IONESCU, "2025-05-23", 4800, None, "6040"),
    (IONESCU, "2025-06-20", 4200, None, "6040"),
    (IONESCU, "2025-07-24", 2200, "2025-08-20", "6040"),

    # Cobalt Structural Engineering PLLC -- occasional, tied to the Whitfield
    # hospitality structural scope and one later Larkspur review
    (COBALT_STRUCTURAL, "2024-09-12", 9500, "2024-10-21", "6040"),
    (COBALT_STRUCTURAL, "2024-10-17", 10200, "2024-11-25", "6040"),
    (COBALT_STRUCTURAL, "2024-11-14", 12500, "2024-12-30", "6040"),
    (COBALT_STRUCTURAL, "2024-12-20", 7800, "2025-02-05", "6040"),
    (COBALT_STRUCTURAL, "2025-02-11", 8000, "2025-03-20", "6040"),
    (COBALT_STRUCTURAL, "2025-05-16", 9200, None, "6040"),

    # Rendercraft Visualization Studio -- occasional, tied to concept/DD
    # presentation milestones
    (RENDERCRAFT, "2024-07-15", 3200, "2024-08-12", "6040"),
    (RENDERCRAFT, "2024-08-14", 3800, "2024-09-11", "6040"),
    (RENDERCRAFT, "2024-10-16", 5200, "2024-11-13", "6040"),
    (RENDERCRAFT, "2024-11-13", 4400, "2024-12-11", "6040"),
    (RENDERCRAFT, "2024-12-12", 4800, "2025-01-13", "6040"),
    (RENDERCRAFT, "2025-01-16", 4200, "2025-02-13", "6040"),
    (RENDERCRAFT, "2025-03-14", 4600, None, "6040"),
    (RENDERCRAFT, "2025-06-11", 4200, None, "6040"),

    # Third Rail Code & Egress Consulting -- occasional, tied to CD/CA phases
    (CODE_CONSULTANT, "2024-09-20", 2600, "2024-10-18", "6040"),
    (CODE_CONSULTANT, "2024-11-18", 2400, "2024-12-16", "6040"),
    (CODE_CONSULTANT, "2025-02-14", 2800, "2025-03-14", "6040"),
    (CODE_CONSULTANT, "2025-04-16", 2200, None, "6040"),

    # Bellweather Procurement & FF&E Sourcing -- occasional, coordinates
    # furniture/fixture/equipment sourcing on the larger hospitality and
    # retail projects (Whitfield, Bellcourt, Larkspur)
    (BELLWEATHER, "2024-10-22", 14000, "2024-11-19", "6040"),
    (BELLWEATHER, "2024-11-26", 17000, "2024-12-23", "6040"),
    (BELLWEATHER, "2025-01-21", 13500, "2025-02-18", "6040"),
    (BELLWEATHER, "2025-03-25", 14000, "2025-04-22", "6040"),
    (BELLWEATHER, "2025-05-19", 13500, None, "6040"),

    # Gramercy Print & Copy -- reprographics and large-format plotting
    (GRAMERCY_PRINT, "2024-09-24", 900, "2024-10-20", "6050"),
    (GRAMERCY_PRINT, "2024-11-21", 1400, "2024-12-18", "6050"),
    (GRAMERCY_PRINT, "2025-01-23", 1100, "2025-02-19", "6050"),
    (GRAMERCY_PRINT, "2025-03-24", 1600, "2025-04-21", "6050"),
    (GRAMERCY_PRINT, "2025-05-27", 1200, None, "6050"),
    (GRAMERCY_PRINT, "2025-06-20", 1400, None, "6050"),

    # DraftLine Software Solutions -- annual CAD/rendering subscription plan,
    # paid up front and expensed as incurred (no prepaid-asset treatment),
    # plus a mid-year second-seat add-on
    (SOFTWARE_VENDOR, "2024-07-08", 8500, "2024-07-25", "6050"),
    (SOFTWARE_VENDOR, "2025-01-08", 3200, "2025-01-25", "6050"),

    # Metro Office Equipment Co -- large-format plotter purchase, expensed
    # as incurred (below the firm's capitalization threshold)
    (EQUIPMENT_VENDOR, "2024-10-08", 4200, "2024-11-05", "6140"),

    # Fillmore Bookkeeping & Tax LLC -- quarterly bookkeeping, plus a one-time
    # outside review of the prior fiscal year's closing books
    (BOOKKEEPER, "2024-09-30", 950, "2024-10-25", "6070"),
    (BOOKKEEPER, "2024-12-31", 1000, "2025-01-20", "6070"),
    (BOOKKEEPER, "2025-01-15", 2600, "2025-02-12", "6070"),
    (BOOKKEEPER, "2025-03-31", 1000, "2025-04-22", "6070"),
    (BOOKKEEPER, "2025-06-30", 1050, None, "6070"),

    # Flatiron Design Quarterly -- print advertising placements
    (AD_VENDOR, "2024-09-16", 650, "2024-09-16", "6130"),
    (AD_VENDOR, "2025-03-11", 700, "2025-03-11", "6130"),
]


def bill_html(vendor: str, addr: str, bill_num: str, issue_display: str, amount: int,
              description: str) -> str:
    return f"""
<div class="header">
<h1>{vendor}</h1>
<div class="small">{addr}</div>
</div>
<h2>Invoice {bill_num}</h2>
<p>To: {FIRM_NAME}<br/>{FIRM_ADDR}<br/>Date: {issue_display}<br/>Terms: Net 30</p>
<table>
<tr><th>Description</th><th class="right">Amount</th></tr>
<tr><td>{description}</td><td class="right">{money_str(amount)}</td></tr>
<tr class="total-row"><td>Total Due</td><td class="right">{money_str(amount)}</td></tr>
</table>
"""


BILLS_BY_MONTH: dict[str, list[tuple]] = {}
for _b in BILLS:
    BILLS_BY_MONTH.setdefault(_b[1][:7], []).append(_b)

# One designated month's bundle is scanned image-only (in addition to being
# a bundle) to preserve genuine scanned-PDF format variety in the corpus --
# a subcontractor forwarding a stack of paperwork as a single scan is exactly
# how this would arrive in real life.
SCANNED_BUNDLE_MONTH = "2024-11"

_bill_variant_counter = 0
for _month_key in sorted(BILLS_BY_MONTH):
    _batch_key = f"monthly-{_month_key}"
    if _batch_key not in BATCH_DIR:
        continue
    _month_bills = sorted(BILLS_BY_MONTH[_month_key], key=lambda b: b[1])
    _tmp_pdfs: list[str] = []
    _bill_meta: list[tuple] = []
    for vendor, issue_date, amt_dollars, paid_date, code in _month_bills:
        amount = cents(amt_dollars)
        variant = _bill_variant_counter % 3
        _bill_variant_counter += 1
        issue_display = fmt_date_variant(issue_date, variant)
        bill_num = f"{re.sub(r'[^A-Za-z]', '', vendor)[:3].upper()}-{issue_date.replace('-', '')}"
        html_content = bill_html(vendor, VENDOR_ADDR[vendor], bill_num, issue_display, amount,
                                  VENDOR_DESC[vendor])
        tmp_pdf = os.path.join(TMP_DIR, f"BILLPG-{bill_num}-{_bill_variant_counter}.pdf")
        R.render_html_to_pdf(html_wrap(html_content), tmp_pdf, css=BASE_CSS)
        _tmp_pdfs.append(tmp_pdf)
        _bill_meta.append((vendor, issue_date, amount, paid_date, code, bill_num, issue_display))

    if len(_tmp_pdfs) > 1:
        fname = f"vendor_bills_{_month_key}.pdf"
        _merged_tmp = os.path.join(TMP_DIR, f"BUNDLE-{_month_key}.pdf")
        R.concat_pdfs(_tmp_pdfs, _merged_tmp)
    else:
        fname = f"vendor_bill_{_month_key}.pdf"
        _merged_tmp = _tmp_pdfs[0]

    out_path = batch_path(_batch_key, fname)
    if len(_bill_meta) > 1 and _month_key == SCANNED_BUNDLE_MONTH:
        R.scanify(_merged_tmp, out_path, seed=SEED + 700)
        _scanned = True
    else:
        shutil.copyfile(_merged_tmp, out_path)
        _scanned = False

    doc_id = next_id("DOC-BILL")
    rel = batch_relpath(_batch_key, fname)
    last_display = _bill_meta[-1][6]
    if len(_bill_meta) == 1:
        kind = "bill_in"
        counterparty_label = _bill_meta[0][0]
        amount_field = _bill_meta[0][2]
    else:
        kind = "multi_document_bundle"
        counterparty_label = "Multiple vendors (see attached)"
        amount_field = None
    add_doc(doc_id, kind, rel, "pdf", _scanned, last_display, counterparty_label, amount_field)

    if len(_bill_meta) == 1:
        v0 = _bill_meta[0][0]
        batch_attach(_batch_key, fname, f"bill from {v0} - {money_str(_bill_meta[0][2])}")
    else:
        _names = ", ".join(sorted(set(b[0] for b in _bill_meta)))
        batch_attach(_batch_key, fname,
                     f"{len(_bill_meta)} vendor bills this month, one file ({_names})")

    for vendor, issue_date, amount, paid_date, code, bill_num, issue_display in _bill_meta:
        eid = next_id("BILL")
        add_entry(eid, issue_date, [
            (code, amount, 0, f"Bill {bill_num} - {vendor}", vendor, [doc_id]),
            ("2000", 0, amount, f"Bill {bill_num} - {vendor}", vendor, [doc_id]),
        ])
        if paid_date:
            peid = next_id("BILLPMT")
            sd = stmt_doc_id(paid_date)
            add_entry(peid, paid_date, [
                ("2000", amount, 0, f"Payment - bill {bill_num} - {vendor}", vendor, [doc_id]),
                ("1000", 0, amount, f"Payment - bill {bill_num} - {vendor}", vendor, [sd]),
            ])

print(f"Bills built: {len(BILLS)} across {len(BILLS_BY_MONTH)} monthly attachments", file=sys.stderr)

# ---------------------------------------------------------------------------
# Fixed / routine monthly cash expenses -- direct debit, evidenced by the
# bank statement itself (no separate vendor invoice mailed to the firm), but
# varying the way a real recurring cost actually varies: a rent step at the
# lease anniversary, a seasonal utility swing, and two separate annual
# insurance policies renewing at different premiums.
# ---------------------------------------------------------------------------

RENT_ESCALATION_MONTH = "2025-01"  # lease anniversary -- rent steps up here
RENT_BEFORE = cents(5500)
RENT_AFTER = cents(5750)

# Telephone & internet: a fixed-rate plan ($180 base) still varies in real
# life -- monthly taxes/surcharges, a provider rate increase at contract
# renewal, and one month with a usage overage.
TELECOM_BASE_BEFORE = cents(180)
TELECOM_BASE_AFTER = cents(188)  # Gotham Fiber & Voice's spring rate increase
TELECOM_RATE_CHANGE_MONTH = "2025-04"
TELECOM_OVERAGE_MONTH = "2024-11"
TELECOM_OVERAGE_AMOUNT = cents(22)

# Seasonal utility swing: higher in winter (heat) and mid-summer (AC),
# lower in the shoulder months -- a real NYC office utility bill, not a
# flat constant.
UTILITY_SEASON_BASE = {
    "07": 520, "08": 560, "09": 420, "10": 380, "11": 460, "12": 610,
    "01": 640, "02": 590, "03": 470, "04": 400, "05": 410, "06": 500,
}

for m in MONTHS:
    rent_amt = RENT_BEFORE if m < RENT_ESCALATION_MONTH else RENT_AFTER
    d1 = f"{m}-03"
    sd = stmt_doc_id(d1)
    eid = next_id("RENT")
    add_entry(eid, d1, [
        ("6000", rent_amt, 0, f"Rent - {MONTH_LABEL[m]} - {LANDLORD}", LANDLORD, [sd]),
        ("1000", 0, rent_amt, f"Rent - {MONTH_LABEL[m]} - {LANDLORD}", LANDLORD, [sd]),
    ])

    base = UTILITY_SEASON_BASE[m[5:7]]
    util_amt = cents(base + RNG.randint(-30, 40))
    d2 = f"{m}-07"
    sd2 = stmt_doc_id(d2)
    eid = next_id("UTIL")
    add_entry(eid, d2, [
        ("6010", util_amt, 0, f"Utilities - {MONTH_LABEL[m]} - {UTILITY_CO}", UTILITY_CO, [sd2]),
        ("1000", 0, util_amt, f"Utilities - {MONTH_LABEL[m]} - {UTILITY_CO}", UTILITY_CO, [sd2]),
    ])

    d3 = f"{m}-05"
    sd3 = stmt_doc_id(d3)
    eid = next_id("TEL")
    telecom_base = TELECOM_BASE_BEFORE if m < TELECOM_RATE_CHANGE_MONTH else TELECOM_BASE_AFTER
    telecom_surcharge = RNG.randint(-150, 260)  # a few dollars of tax/surcharge drift, either way
    telecom_amt = telecom_base + telecom_surcharge
    telecom_memo = f"Phone & internet - {MONTH_LABEL[m]} - {TELECOM_CO}"
    if m == TELECOM_OVERAGE_MONTH:
        telecom_amt += TELECOM_OVERAGE_AMOUNT
        telecom_memo = f"Phone & internet (incl. usage overage) - {MONTH_LABEL[m]} - {TELECOM_CO}"
    add_entry(eid, d3, [
        ("6150", telecom_amt, 0, telecom_memo, TELECOM_CO, [sd3]),
        ("1000", 0, telecom_amt, telecom_memo, TELECOM_CO, [sd3]),
    ])

    bank_fee = cents(RNG.randint(16, 26))
    d4 = f"{m}-01"
    sd4 = stmt_doc_id(d4)
    eid = next_id("BFEE")
    add_entry(eid, d4, [
        ("6080", bank_fee, 0, f"Monthly account fee - {MONTH_LABEL[m]} - {BANK_NAME}", BANK_NAME, [sd4]),
        ("1000", 0, bank_fee, f"Monthly account fee - {MONTH_LABEL[m]} - {BANK_NAME}", BANK_NAME, [sd4]),
    ])

    storage_amt = cents(295)
    d5 = f"{m}-09"
    sd5 = stmt_doc_id(d5)
    eid = next_id("STORAGE")
    add_entry(eid, d5, [
        ("6900", storage_amt, 0, f"Self-storage unit - {MONTH_LABEL[m]} - {STORAGE_VENDOR}", STORAGE_VENDOR, [sd5]),
        ("1000", 0, storage_amt, f"Self-storage unit - {MONTH_LABEL[m]} - {STORAGE_VENDOR}", STORAGE_VENDOR, [sd5]),
    ])

# Insurance -- two separate annual policies (general liability, and
# professional liability / E&O from a different carrier), each renewing once
# in the period at a different premium than the year before.
GL_POLICY = [
    ("2024-07-12", 3200),
    ("2025-07-12", 3450),  # renewal, trailing stub
]
EO_POLICY = [
    ("2025-01-18", 4300),
    # prior E&O term (2024-01-18 to 2025-01-17) was paid before this period
    # began and is not repeated here.
]
for d, amt_dollars in GL_POLICY:
    sd = stmt_doc_id(d)
    eid = next_id("INS")
    amt = cents(amt_dollars)
    add_entry(eid, d, [
        ("6060", amt, 0, f"Annual general liability premium - {INSURER}", INSURER, [sd]),
        ("1000", 0, amt, f"Annual general liability premium - {INSURER}", INSURER, [sd]),
    ])
for d, amt_dollars in EO_POLICY:
    sd = stmt_doc_id(d)
    eid = next_id("INS")
    amt = cents(amt_dollars)
    add_entry(eid, d, [
        ("6060", amt, 0, f"Annual professional liability (E&O) premium - {SECOND_INSURER}", SECOND_INSURER, [sd]),
        ("1000", 0, amt, f"Annual professional liability (E&O) premium - {SECOND_INSURER}", SECOND_INSURER, [sd]),
    ])

print("Fixed monthly expenses built.", file=sys.stderr)

# ---------------------------------------------------------------------------
# Payroll: one salaried employee, monthly-in-arrears via a payroll provider.
# ---------------------------------------------------------------------------

GROSS_MONTHLY = cents(6000)
EMP_TAX_MONTHLY = cents(480)
TOTAL_DRAFT = GROSS_MONTHLY + EMP_TAX_MONTHLY

# work_month -> pay_date, for normal (non-accrual) pay runs
PAYROLL_RUNS = [
    ("2024-07", "2024-08-05"), ("2024-08", "2024-09-05"), ("2024-09", "2024-10-05"),
    ("2024-10", "2024-11-05"), ("2024-11", "2024-12-05"), ("2024-12", "2025-01-05"),
    ("2025-01", "2025-02-05"), ("2025-02", "2025-03-05"), ("2025-03", "2025-04-05"),
    ("2025-04", "2025-05-05"), ("2025-05", "2025-06-05"),
    # 2025-06 work is accrued at period end, settled 2025-07-05 (below)
    ("2025-07", "2025-08-05"),
]

DOC_PR_ACCRUAL = "DOC-PAYROLL-ACCRUAL"


def payroll_summary_html(title: str, rows: list[tuple[str, str, int, int, int]]) -> str:
    body = f"""
<div class="header">
<h1>{PAYROLL_PROVIDER}</h1>
<div class="small">{PAYROLL_PROVIDER_ADDR}</div>
</div>
<h2>{title}</h2>
<p>Client: {FIRM_NAME}<br/>Employee: {EMPLOYEE_NAME} ({EMPLOYEE_TITLE})</p>
<table>
<tr><th>Pay Period</th><th>Pay Date</th><th class="right">Gross Wages</th><th class="right">Employer Taxes</th><th class="right">Total Drafted</th></tr>
"""
    for period, pay_date, gross, tax, total in rows:
        body += f"<tr><td>{period}</td><td>{pay_date}</td><td class='right'>{money_str(gross)}</td><td class='right'>{money_str(tax)}</td><td class='right'>{money_str(total)}</td></tr>\n"
    body += "</table><p class='small'>Amounts drafted from the client's operating account cover gross wages, employer payroll taxes, and employee withholding remitted on the employee's behalf.</p>"
    return body


def build_payroll_bundle(key: str, title: str, filename: str, doc_id: str, issue_date: str,
                          runs_slice: list[tuple[str, str]]) -> None:
    rows = []
    for work_month, pay_date in runs_slice:
        period_label = f"{MONTH_LABEL[work_month]} (worked)"
        rows.append((period_label, pay_date, GROSS_MONTHLY, EMP_TAX_MONTHLY, TOTAL_DRAFT))
    html = payroll_summary_html(title, rows)
    render_pdf(key, filename, html, "payroll_summary", fmt_date_variant(issue_date, 1), PAYROLL_PROVIDER, None,
               doc_id=doc_id)
    n = len(runs_slice)
    batch_attach(key, filename, f"payroll summary from {PAYROLL_PROVIDER} covering {n} pay run{'s' if n != 1 else ''}")
    for work_month, pay_date in runs_slice:
        eid = next_id("PAYROLL")
        sd = stmt_doc_id(pay_date)
        add_entry(eid, pay_date, [
            ("6020", GROSS_MONTHLY, 0, f"Salary - {EMPLOYEE_NAME} - {MONTH_LABEL[work_month]}", EMPLOYEE_NAME, [doc_id, sd]),
            ("6030", EMP_TAX_MONTHLY, 0, f"Employer payroll tax - {EMPLOYEE_NAME} - {MONTH_LABEL[work_month]}", EMPLOYEE_NAME, [doc_id, sd]),
            ("1000", 0, TOTAL_DRAFT, f"Payroll draft - {EMPLOYEE_NAME} - {MONTH_LABEL[work_month]}", PAYROLL_PROVIDER, [doc_id, sd]),
        ])


build_payroll_bundle("monthly-2024-10", "Payroll Summary - Q1 FY2025 Pay Runs", "payroll_summary_q1.pdf",
                     next_id("DOC-PAYROLL"), "2024-10-05", PAYROLL_RUNS[0:3])
build_payroll_bundle("monthly-2025-01", "Payroll Summary - Q2 FY2025 Pay Runs", "payroll_summary_q2.pdf",
                     next_id("DOC-PAYROLL"), "2025-01-05", PAYROLL_RUNS[3:6])
build_payroll_bundle("monthly-2025-04", "Payroll Summary - Q3 FY2025 Pay Runs", "payroll_summary_q3.pdf",
                     next_id("DOC-PAYROLL"), "2025-04-05", PAYROLL_RUNS[6:9])
build_payroll_bundle("monthly-2025-06", "Payroll Summary - Q4 FY2025 Pay Runs (through May)", "payroll_summary_q4.pdf",
                     next_id("DOC-PAYROLL"), "2025-06-05", PAYROLL_RUNS[9:11])
build_payroll_bundle("monthly-2025-08", "Payroll Summary - Pay Runs Through August 2025", "payroll_summary_q5.pdf",
                     next_id("DOC-PAYROLL"), "2025-08-05", PAYROLL_RUNS[11:12])

# June 2025 accrual: work performed by period end, pay date 2025-07-05 (after
# period_end) -- this is the accrued payroll liability required by the brief.
accrual_html = f"""
<div class="header"><h1>{PAYROLL_PROVIDER}</h1><div class="small">{PAYROLL_PROVIDER_ADDR}</div></div>
<h2>Upcoming Pay Run Notice</h2>
<p>Client: {FIRM_NAME}<br/>Employee: {EMPLOYEE_NAME} ({EMPLOYEE_TITLE})</p>
<p>This confirms the pay run scheduled to draft on 2025-07-05 for hours worked
in June 2025:</p>
<table>
<tr><th>Pay Period</th><th>Scheduled Pay Date</th><th class="right">Gross Wages</th><th class="right">Employer Taxes</th><th class="right">Total to be Drafted</th></tr>
<tr><td>June 2025 (worked)</td><td>2025-07-05</td><td class="right">{money_str(GROSS_MONTHLY)}</td><td class="right">{money_str(EMP_TAX_MONTHLY)}</td><td class="right">{money_str(TOTAL_DRAFT)}</td></tr>
</table>
"""
render_pdf("monthly-2025-06", "payroll_accrual_notice_june.pdf", accrual_html, "payroll_summary",
           "2025-06-25", PAYROLL_PROVIDER, None, doc_id=DOC_PR_ACCRUAL)
batch_attach("monthly-2025-06", "payroll_accrual_notice_june.pdf",
             f"payroll notice from {PAYROLL_PROVIDER} - June work, paid July 5 (accrued at fiscal year end)")

add_entry(next_id("PAYROLL-ACCR"), "2025-06-30", [
    ("6020", GROSS_MONTHLY, 0, f"Accrued salary - {EMPLOYEE_NAME} - {MONTH_LABEL['2025-06']} (paid 2025-07-05)", EMPLOYEE_NAME, [DOC_PR_ACCRUAL]),
    ("6030", EMP_TAX_MONTHLY, 0, f"Accrued employer payroll tax - {EMPLOYEE_NAME} - {MONTH_LABEL['2025-06']}", EMPLOYEE_NAME, [DOC_PR_ACCRUAL]),
    ("2200", 0, TOTAL_DRAFT, f"Accrued payroll liability - {EMPLOYEE_NAME} - {MONTH_LABEL['2025-06']}", EMPLOYEE_NAME, [DOC_PR_ACCRUAL]),
])

sd = stmt_doc_id("2025-07-05")
add_entry(next_id("PAYROLL-SETTLE"), "2025-07-05", [
    ("2200", TOTAL_DRAFT, 0, f"Settle accrued payroll - {EMPLOYEE_NAME} - {MONTH_LABEL['2025-06']}", EMPLOYEE_NAME, [DOC_PR_ACCRUAL]),
    ("1000", 0, TOTAL_DRAFT, f"Settle accrued payroll - {EMPLOYEE_NAME} - {MONTH_LABEL['2025-06']}", PAYROLL_PROVIDER, [sd]),
])

print("Payroll built.", file=sys.stderr)

# ---------------------------------------------------------------------------
# Partner draws -- recurring monthly withdrawals, different per partner.
# ---------------------------------------------------------------------------

DRAW_HALLORAN = cents(6500)
DRAW_VANCE = cents(4800)
ALL_DRAW_MONTHS = MONTHS  # Jul2024 - Aug2025, FY + trailing stub

for m in ALL_DRAW_MONTHS:
    d = f"{m}-28"
    sd = stmt_doc_id(d)
    eid = next_id("DRAW-H")
    add_entry(eid, d, [
        ("3120", DRAW_HALLORAN, 0, f"Partner draw - Halloran - {MONTH_LABEL[m]}", HALLORAN_NAME, [sd]),
        ("1000", 0, DRAW_HALLORAN, f"Partner draw - Halloran - {MONTH_LABEL[m]}", HALLORAN_NAME, [sd]),
    ])
    eid = next_id("DRAW-V")
    add_entry(eid, d, [
        ("3130", DRAW_VANCE, 0, f"Partner draw - Vance - {MONTH_LABEL[m]}", VANCE_NAME, [sd]),
        ("1000", 0, DRAW_VANCE, f"Partner draw - Vance - {MONTH_LABEL[m]}", VANCE_NAME, [sd]),
    ])

# ---------------------------------------------------------------------------
# Mandated defect: personal expense on the business account, treated as a
# partner draw (not an expense).
# ---------------------------------------------------------------------------

personal_date = "2025-02-14"
personal_amount = cents(612)
receipt_doc = render_pdf("monthly-2025-02", "atlanticcrest_receipt_personal.pdf", f"""
<div class="header"><h1>Atlantic Crest Airlines</h1><div class="small">E-Receipt / Passenger Itinerary</div></div>
<p>Passenger: M. Halloran<br/>Date of travel: {fmt_date_variant(personal_date, 1)}<br/>
Route: New York (LGA) - Miami (MIA) - New York (LGA)<br/>Fare class: Main Cabin</p>
<table><tr><th>Description</th><th class="right">Amount</th></tr>
<tr><td>Round-trip airfare</td><td class="right">{money_str(personal_amount)}</td></tr>
<tr class="total-row"><td>Total Charged</td><td class="right">{money_str(personal_amount)}</td></tr></table>
<p class="small">Charged to card ending 2289.</p>
""", "receipt", fmt_date_variant(personal_date, 1), "Atlantic Crest Airlines", personal_amount)
batch_attach("monthly-2025-02", "atlanticcrest_receipt_personal.pdf",
             f"Atlantic Crest Airlines receipt, {money_str(personal_amount)} - personal weekend trip, not project travel")

sd = stmt_doc_id(personal_date)
_personal_eid = next_id("PERSONAL")
BANK_DESC[_personal_eid] = "DEBIT CARD PURCHASE - ATLANTIC CREST AIR 8842"
add_entry(_personal_eid, personal_date, [
    ("3120", personal_amount, 0, "Personal airfare (M. Halloran, Miami weekend) charged to business account -- reclassified as partner draw, not a business expense", HALLORAN_NAME, [sd, receipt_doc]),
    ("1000", 0, personal_amount, "Personal airfare (M. Halloran, Miami weekend) charged to business account -- reclassified as partner draw, not a business expense", HALLORAN_NAME, [sd, receipt_doc]),
])
batch_para("monthly-2025-02",
    f"One more thing -- the Atlantic Crest charge on the card this month "
    f"({money_str(personal_amount)}) was our Miami trip, not a client visit. "
    f"Please book that against my draw account, not travel expense.")

# ---------------------------------------------------------------------------
# Mandated: one partner contributes additional capital once in the period.
# ---------------------------------------------------------------------------

capital_date = "2024-12-02"
capital_amount = cents(45000)
note_doc = render_txt_doc("monthly-2024-12", "vance_capital_contribution_note.txt", f"""{VANCE_NAME}
{FIRM_NAME}

{fmt_date_variant(capital_date, 1)}

Note for the file: wiring an additional {money_str(capital_amount)} of my own
funds into the operating account to cover the cash gap this month -- the
Whitfield Phase 3 milestone invoice went out December 12 but won't be
collected until January, and the subcontractor bills tied to that project
are due now. This is a capital contribution, not a loan -- please credit it
to my partner capital account.

-- Owen
""", "capital_contribution_note", fmt_date_variant(capital_date, 1), VANCE_NAME, capital_amount)
batch_attach("monthly-2024-12", "vance_capital_contribution_note.txt",
             f"note from Owen Vance - additional capital contribution, {money_str(capital_amount)}")

sd = stmt_doc_id(capital_date)
add_entry(next_id("CAPCONTRIB"), capital_date, [
    ("1000", capital_amount, 0, "Additional capital contribution - Vance", VANCE_NAME, [sd, note_doc]),
    ("3110", 0, capital_amount, "Additional capital contribution - Vance", VANCE_NAME, [sd, note_doc]),
])
batch_para("monthly-2024-12",
    f"Also attaching a short note from Owen -- he wired {money_str(capital_amount)} into the "
    f"operating account on {fmt_date_variant(capital_date, 1)}. That's an additional capital "
    f"contribution on his side, not a loan to the firm, so please book it to his capital account.")

print("Draws, personal expense, and capital contribution built.", file=sys.stderr)

# ---------------------------------------------------------------------------
# Mandated defect: duplicate receipt -- same purchase shipped twice, two
# different formats (text PDF emailed directly, and a photographed JPG
# attached later to the self-maintained expense log).
# ---------------------------------------------------------------------------

dup_date = "2024-09-11"
dup_amount = cents(215)

dup_pdf_doc = render_pdf("monthly-2024-09", "tribeca_paper_receipt.pdf", f"""
<div class="header"><h1>{OFFICE_SUPPLY_VENDOR}</h1><div class="small">214 Church Street, New York, NY 10013</div></div>
<h2>Sales Receipt</h2>
<p>Date: {fmt_date_variant(dup_date, 0)}<br/>Sold to: {FIRM_NAME}</p>
<table><tr><th>Item</th><th class="right">Amount</th></tr>
<tr><td>Presentation boards, drafting paper, print supplies</td><td class="right">{money_str(dup_amount)}</td></tr>
<tr class="total-row"><td>Total</td><td class="right">{money_str(dup_amount)}</td></tr></table>
""", "receipt", fmt_date_variant(dup_date, 0), OFFICE_SUPPLY_VENDOR, dup_amount)
batch_attach("monthly-2024-09", "tribeca_paper_receipt.pdf",
             f"receipt from {OFFICE_SUPPLY_VENDOR} - {money_str(dup_amount)} - office/print supplies")

sd = stmt_doc_id(dup_date)
add_entry(next_id("SUPPLIES"), dup_date, [
    ("6050", dup_amount, 0, f"Office/print supplies - {OFFICE_SUPPLY_VENDOR}", OFFICE_SUPPLY_VENDOR, [sd, dup_pdf_doc]),
    ("1000", 0, dup_amount, f"Office/print supplies - {OFFICE_SUPPLY_VENDOR}", OFFICE_SUPPLY_VENDOR, [sd, dup_pdf_doc]),
])
# doc_ids for the expense line get the duplicate jpg + xlsx log added below,
# once those documents exist -- but the ledger already cites the PDF only,
# by design: there is exactly one real expense, and the SECOND shipped copy
# (photographed jpg, in the October expense log) is never referenced by the
# ledger. It exists in the corpus purely to demonstrate the double-counting
# trap for a naive ingestion pass; see answer-key.md defect #1.

# ---------------------------------------------------------------------------
# Mandated defect: one handwritten-looking cash receipt, photographed at an
# angle.
# ---------------------------------------------------------------------------

hw_date = "2025-04-11"
hw_amount = cents(84)
hw_doc = render_handwritten_jpg("monthly-2025-04", "chelsea_hardware_receipt.jpg", [
    "CHELSEA HARDWARE",
    "228 9th Ave",
    fmt_date_variant(hw_date, 2),
    "hardware",
    f"TOTAL {money_str(hw_amount)}",
    "CASH",
], "cash_receipt_handwritten", fmt_date_variant(hw_date, 2), HARDWARE_VENDOR, hw_amount, seed=SEED + 1)
batch_attach("monthly-2025-04", "chelsea_hardware_receipt.jpg",
             f"handwritten receipt from {HARDWARE_VENDOR} - {money_str(hw_amount)} - hardware for the Pemberton punch list")

sd = stmt_doc_id(hw_date)
add_entry(next_id("REPAIRS"), hw_date, [
    ("6140", hw_amount, 0, f"Hardware/supplies - {HARDWARE_VENDOR}", HARDWARE_VENDOR, [sd, hw_doc]),
    ("1000", 0, hw_amount, f"Hardware/supplies - {HARDWARE_VENDOR}", HARDWARE_VENDOR, [sd, hw_doc]),
])

print("Duplicate receipt and handwritten receipt built.", file=sys.stderr)

# ---------------------------------------------------------------------------
# Self-maintained expense-log months: small ad hoc expenses that arrive only
# as an XLSX log with receipts attached separately (no individual bill_in
# per item). The September Tribeca purchase reappears here as the duplicate
# (photographed jpg), logged alongside genuinely log-only items.
# ---------------------------------------------------------------------------

EXPENSE_LOGS = [
    {
        "key": "expense-log-4", "month_label": "July 2024", "issue_date": "2024-08-14",
        "items": [
            {"date": "2024-07-09", "vendor": FRAME_VENDOR, "desc": "Framed presentation boards - Ansel DD phase 2",
             "amount": cents(265), "code": "6050", "receipt": True, "seed_key": "log4a"},
            {"date": "2024-07-16", "vendor": "Metro Car Service", "desc": "Site visit - Ansel Family Residence",
             "amount": cents(145), "code": "6900", "receipt": False, "seed_key": None},
            {"date": "2024-07-24", "vendor": "Corner Deli Catering", "desc": "Client meeting lunch - Pemberton kickoff",
             "amount": cents(88), "code": "6900", "receipt": False, "seed_key": None},
            {"date": "2024-07-26", "vendor": "Chelsea Hardware & Supply", "desc": "Studio equipment - cables and drafting supplies",
             "amount": cents(210), "code": "6900", "receipt": False, "seed_key": None},
            {"date": "2024-07-30", "vendor": COURIER_VENDOR, "desc": "Drawing set delivery - Pemberton Townhouse",
             "amount": cents(62), "code": "6900", "receipt": False, "seed_key": None},
        ],
    },
    {
        "key": "expense-log-1", "month_label": "September 2024", "issue_date": "2024-10-08",
        "items": [
            {"date": "2024-09-11", "vendor": OFFICE_SUPPLY_VENDOR, "desc": "Presentation boards, drafting paper, print supplies (see attached photo)",
             "amount": dup_amount, "code": "6050", "receipt": True, "seed_key": "dup"},
            {"date": "2024-09-23", "vendor": GRAMERCY_PRINT, "desc": "Large-format plots for client presentation",
             "amount": cents(96), "code": "6050", "receipt": False, "seed_key": None},
            {"date": "2024-09-27", "vendor": "MTA Metrocard", "desc": "Site-visit transit fare reimbursement",
             "amount": cents(33), "code": "6900", "receipt": False, "seed_key": None},
            {"date": "2024-09-19", "vendor": "NoMad Design Materials Library", "desc": "Fabric and finish samples - Bellcourt retail",
             "amount": cents(340), "code": "6050", "receipt": False, "seed_key": None},
            {"date": "2024-09-25", "vendor": COURIER_VENDOR, "desc": "Drawing delivery - Bellcourt Retail Group site",
             "amount": cents(58), "code": "6900", "receipt": False, "seed_key": None},
            {"date": "2024-09-30", "vendor": DUES_ORG, "desc": "Annual firm membership dues",
             "amount": cents(450), "code": "6070", "receipt": False, "seed_key": None},
        ],
    },
    {
        "key": "expense-log-5", "month_label": "November 2024", "issue_date": "2024-12-18",
        "items": [
            {"date": "2024-11-06", "vendor": "NoMad Design Materials Library", "desc": "Fabric and finish samples - Whitfield closeout",
             "amount": cents(610), "code": "6050", "receipt": False, "seed_key": None},
            {"date": "2024-11-12", "vendor": "Metro Car Service", "desc": "Site visits - multiple trips, Whitfield & Larkspur",
             "amount": cents(260), "code": "6900", "receipt": False, "seed_key": None},
            {"date": "2024-11-20", "vendor": "Old Fulton Chophouse", "desc": "Client dinner - Whitfield walkthrough",
             "amount": cents(210), "code": "6900", "receipt": False, "seed_key": None},
            {"date": "2024-11-25", "vendor": COURIER_VENDOR, "desc": "Blueprint delivery, two trips",
             "amount": cents(95), "code": "6900", "receipt": False, "seed_key": None},
            {"date": "2024-11-27", "vendor": GRAMERCY_PRINT, "desc": "Additional plots outside monthly account",
             "amount": cents(180), "code": "6050", "receipt": False, "seed_key": None},
        ],
    },
    {
        "key": "expense-log-2", "month_label": "January 2025", "issue_date": "2025-02-06",
        "items": [
            {"date": "2025-01-14", "vendor": FRAME_VENDOR, "desc": "Framed renderings for Whitfield presentation",
             "amount": cents(178), "code": "6050", "receipt": True, "seed_key": "log2a"},
            {"date": "2025-01-22", "vendor": OFFICE_SUPPLY_VENDOR, "desc": "Print supplies restock",
             "amount": cents(142), "code": "6050", "receipt": False, "seed_key": None},
            {"date": "2025-01-30", "vendor": "Corner Deli Catering", "desc": "Client meeting lunch - Meridian kickoff",
             "amount": cents(64), "code": "6900", "receipt": False, "seed_key": None},
            {"date": "2025-01-09", "vendor": DUES_ORG, "desc": "NCIDQ continuing-education course",
             "amount": cents(525), "code": "6070", "receipt": False, "seed_key": None},
            {"date": "2025-01-27", "vendor": "Metro-North Railroad", "desc": "Site visit + overnight - Larkspur Hospitality (Hudson Valley)",
             "amount": cents(780), "code": "6900", "receipt": False, "seed_key": None},
            {"date": "2025-01-20", "vendor": "Chelsea Hardware & Supply", "desc": "Replacement studio monitor",
             "amount": cents(340), "code": "6900", "receipt": False, "seed_key": None},
        ],
    },
    {
        "key": "expense-log-6", "month_label": "February 2025", "issue_date": "2025-03-14",
        "items": [
            {"date": "2025-02-07", "vendor": "NoMad Design Materials Library", "desc": "Material library restock",
             "amount": cents(410), "code": "6050", "receipt": False, "seed_key": None},
            {"date": "2025-02-13", "vendor": "Corner Deli Catering", "desc": "Client lunches - Ansel and Meridian",
             "amount": cents(150), "code": "6900", "receipt": False, "seed_key": None},
            {"date": "2025-02-21", "vendor": COURIER_VENDOR, "desc": "Sample delivery - Ansel Family Residence",
             "amount": cents(70), "code": "6900", "receipt": False, "seed_key": None},
            {"date": "2025-02-25", "vendor": "Metro Car Service", "desc": "Site visits - Ansel and Meridian",
             "amount": cents(310), "code": "6900", "receipt": False, "seed_key": None},
        ],
    },
    {
        "key": "expense-log-3", "month_label": "April 2025", "issue_date": "2025-05-09",
        "items": [
            {"date": "2025-04-08", "vendor": "Chelsea Hardware & Supply", "desc": "Sample-board mounting hardware",
             "amount": cents(47), "code": "6050", "receipt": False, "seed_key": None},
            {"date": "2025-04-21", "vendor": GRAMERCY_PRINT, "desc": "Bound spec books, 3 copies",
             "amount": cents(210), "code": "6050", "receipt": True, "seed_key": "log3a"},
            {"date": "2025-04-29", "vendor": "MTA Metrocard", "desc": "Site-visit transit fare reimbursement",
             "amount": cents(28), "code": "6900", "receipt": False, "seed_key": None},
            {"date": "2025-04-11", "vendor": DUES_ORG, "desc": "Flatiron Design Forum conference registration",
             "amount": cents(395), "code": "6070", "receipt": False, "seed_key": None},
            {"date": "2025-04-24", "vendor": "Chelsea Hardware & Supply", "desc": "Sample-cabinet hardware",
             "amount": cents(480), "code": "6900", "receipt": False, "seed_key": None},
        ],
    },
    {
        "key": "expense-log-7", "month_label": "June 2025", "issue_date": "2025-07-16",
        "items": [
            {"date": "2025-06-05", "vendor": "NoMad Design Materials Library", "desc": "Closeout sample returns and final materials",
             "amount": cents(290), "code": "6050", "receipt": False, "seed_key": None},
            {"date": "2025-06-13", "vendor": DUES_ORG, "desc": "Annual software/CE license renewal",
             "amount": cents(275), "code": "6070", "receipt": False, "seed_key": None},
            {"date": "2025-06-19", "vendor": "Chelsea Hardware & Supply", "desc": "Tablet stylus and drafting tools",
             "amount": cents(260), "code": "6900", "receipt": False, "seed_key": None},
            {"date": "2025-06-24", "vendor": "Metro Car Service", "desc": "Final site visit - Meridian Law Offices",
             "amount": cents(195), "code": "6900", "receipt": False, "seed_key": None},
            {"date": "2025-06-27", "vendor": COURIER_VENDOR, "desc": "Final drawing delivery - Fenwick Hospitality",
             "amount": cents(65), "code": "6900", "receipt": False, "seed_key": None},
        ],
    },
]

for log in EXPENSE_LOGS:
    key = log["key"]
    rows = [["Date", "Vendor", "Description", "Amount (USD)"]]
    for item in log["items"]:
        rows.append([item["date"], item["vendor"], item["desc"], f"{item['amount']/100:.2f}"])
    xlsx_doc = render_xlsx_doc(key, f"expense_log_{log['issue_date']}.xlsx",
                                {"Expenses": rows}, "expense_log", fmt_date_variant(log["issue_date"], 1),
                                FIRM_NAME, None)
    batch_attach(key, f"expense_log_{log['issue_date']}.xlsx",
                 f"self-maintained expense log for {log['month_label']} ({len(log['items'])} line items)")

    for item in log["items"]:
        doc_ids = [xlsx_doc]
        sd = stmt_doc_id(item["date"])
        doc_ids.append(sd)
        if item["receipt"]:
            vendor_slug = re.sub(r"[^a-z0-9]+", "_", item["vendor"].lower()).strip("_")
            fname = f"receipt_{vendor_slug}_{item['date']}.jpg"
            import zlib
            seed_val = SEED + 2 + (zlib.crc32(item["seed_key"].encode()) % 1000)
            rdoc = render_receipt_jpg(key, fname, [
                item["vendor"], fmt_date_variant(item["date"], 1), item["desc"][:40],
                f"TOTAL  {money_str(item['amount'])}",
            ], "receipt", fmt_date_variant(item["date"], 1), item["vendor"], item["amount"], seed=seed_val)
            batch_attach(key, fname, f"photographed receipt - {item['vendor']} - {money_str(item['amount'])}")
            if item["seed_key"] == "dup":
                # This is the duplicate: the real expense line above already
                # cites the PDF; this jpg is shipped but deliberately NOT
                # cited by any ledger line (see note above).
                pass
            else:
                doc_ids.append(rdoc)
        if item["seed_key"] != "dup":
            eid = next_id("ADHOC")
            add_entry(eid, item["date"], [
                (item["code"], item["amount"], 0, f"{item['desc']} - {item['vendor']}", item["vendor"], doc_ids),
                ("1000", 0, item["amount"], f"{item['desc']} - {item['vendor']}", item["vendor"], doc_ids),
            ])

print("Expense logs built.", file=sys.stderr)

# ---------------------------------------------------------------------------
# No fiscal year-end closing entries. Partner draws (3120/3130) are
# contra-equity, debit-normal, and stand open at period end -- they are
# evidenced entirely by the individual dated bank withdrawals already
# posted above. `lib/ledger.py`'s `balance_sheet_totals` nets each account's
# balance by its account TYPE (not its own `normal_side`), so a debit
# balance in an equity-type contra account is correctly subtracted from
# total equity without any derived closing entry. Closing draws to capital
# would be a current-period-derived computation that no single shipped
# document states -- exactly what Hard Rule One forbids -- so it is not
# done here.
# ---------------------------------------------------------------------------

halloran_draws_ytd = L.account_balance_cents(LEDGER, "3120", as_of=PERIOD_END)
vance_draws_ytd = L.account_balance_cents(LEDGER, "3130", as_of=PERIOD_END)

print(f"Partner draws standing open at period end (Halloran draws {money_str(halloran_draws_ytd)}, "
      f"Vance draws {money_str(vance_draws_ytd)}) -- no closing entries posted.", file=sys.stderr)

# ---------------------------------------------------------------------------
# Statements: derive per-account-month from the completed ledger itself, so
# every cash-touching line is guaranteed to reconcile.
# ---------------------------------------------------------------------------

STATEMENTS: list[dict] = []


def build_statements():
    cash_lines = [l for l in LEDGER if l["account_code"] == "1000"]
    by_month: dict[str, list[dict]] = {}
    for l in cash_lines:
        by_month.setdefault(l["date"][:7], []).append(l)

    running = OPEN_CASH
    for m in MONTHS:
        lines = sorted(by_month.get(m, []), key=lambda l: (l["date"], l["entry_id"]))
        start, end = month_bounds(m)
        opening = running
        stmt_lines = []
        for l in lines:
            if l["debit"]:
                amt, direction = l["debit"], "in"
            else:
                amt, direction = l["credit"], "out"
            stmt_lines.append({
                "date": l["date"], "description": BANK_DESC.get(l["entry_id"], l["memo"]), "amount": amt,
                "direction": direction, "entry_id": l["entry_id"],
            })
            running += amt if direction == "in" else -amt
        closing = running
        stmt_id = f"STMT-1000-{m}"
        STATEMENTS.append({
            "stmt_id": stmt_id, "account_code": "1000", "stmt_period_start": start,
            "stmt_period_end": end, "opening_balance": opening, "closing_balance": closing,
            "doc_ids": [STMT_DOC_ID[m]], "lines": stmt_lines,
        })

        # Render the statement PDF now that we have the real figures.
        rows_html = "".join(
            f"<tr><td>{sl['date']}</td><td>{sl['description']}</td>"
            f"<td class='right'>{'+' if sl['direction']=='in' else '-'}{money_str(sl['amount'])}</td></tr>\n"
            for sl in stmt_lines
        )
        html = f"""
<div class="header">
<h1>{BANK_NAME}</h1>
<div class="small">{BANK_ADDR} &middot; {BANK_PHONE}</div>
</div>
<h2>Business Checking Statement</h2>
<p>Account: {FIRM_NAME} &middot; Account No. {BANK_ACCT_MASK}<br/>
Statement Period: {start} to {end}</p>
<table>
<tr><td>Opening Balance</td><td class="right">{money_str(opening)}</td></tr>
</table>
<table>
<tr><th>Date</th><th>Description</th><th class="right">Amount</th></tr>
{rows_html}
</table>
<table>
<tr class="total-row"><td colspan="2">Closing Balance</td><td class="right">{money_str(closing)}</td></tr>
</table>
"""
        key = STMT_BATCH_FOR_MONTH[m]
        out_path = batch_path(key, STMT_FILENAME[m])
        R.render_html_to_pdf(html_wrap(html), out_path, css=BASE_CSS)
        rel = batch_relpath(key, STMT_FILENAME[m])
        add_doc(STMT_DOC_ID[m], "bank_statement", rel, "pdf", False, end, BANK_NAME, None)
        batch_attach(key, STMT_FILENAME[m], f"{BANK_NAME} statement, {MONTH_LABEL[m]} ({money_str(opening)} -> {money_str(closing)})")


build_statements()
batch_para("stmt-6",
    "This notice covers two statement periods. Paperless delivery for this "
    "account was interrupted in June and has now been restored.")
batch_para("stmt-1",
    "Two statement periods are included in this notice. Our records show the "
    "July notification was not delivered; it is resent here together with August.")

print("Statements built and rendered.", file=sys.stderr)

# ---------------------------------------------------------------------------
# Batch email bodies
# ---------------------------------------------------------------------------

MONTHLY_OPENERS = [
    "Here's this month's batch for the file -- invoices we sent out, and what came in from the subs.",
    "Attaching what we've got for the month: invoices out and the bills that came in.",
    "Month-end drop for the books -- see attached.",
    "Quick one this time -- just the usual invoices and sub bills, attached below.",
    "Everything for the month is attached: invoices out, bills in.",
    "Not much to flag on this one -- the month's invoices and bills are attached as usual.",
    "Attached: this month's invoices and the bills that came in from the subs.",
    "Rounding out the file for the month with the attached invoices and bills.",
    "The usual monthly drop is attached -- invoices out, bills in.",
    "Sending over the month's invoices and bills so the file stays current.",
    "Another month, same routine: invoices out, bills in, both attached.",
    "For the file -- this month's invoices and whatever came in from the subs.",
    "Batch for the month is attached; nothing unusual to flag here.",
    "Getting this month's invoices and bills into the file before it piles up.",
    "This one's on the lighter side -- a few invoices and the sub bills, attached.",
    "Wrapping up the month's paperwork -- see the attached invoices and bills.",
]
MONTHLY_SIGNOFFS = ["Thanks,\nMargaret", "Best,\nMargaret", "Thanks -- Margaret", "-- M"]

# Drawn without replacement (one shuffled pool) so no two monthly emails
# open with the same sentence -- there are 14 monthly batches and 16
# openers above, varied in sentence shape, not just wording.
_monthly_opener_pool = MONTHLY_OPENERS.copy()
RNG.shuffle(_monthly_opener_pool)

STMT_OPENERS = [
    "Your statement is ready to view online. A PDF copy is attached for your records.",
    "Please find attached the statement(s) for your account, as requested.",
]

EXPENSE_LOG_OPENERS = [
    "Sorry for the delay getting this one over -- here's the expense log with receipts attached.",
    "Finally got the receipts together for this month; log and scans attached.",
]


def default_fill_batch(key: str) -> None:
    """Fill in From/To/Subject for batches that never got explicit values."""
    email = BATCH_EMAIL[key]
    if email["from"]:
        return
    if key.startswith("monthly-"):
        idx = list(BATCH_DIR.keys()).index(key)
        email["from"] = f"{HALLORAN_NAME} <{HALLORAN_EMAIL}>"
        email["to"] = f"{VANCE_NAME} <{VANCE_EMAIL}>, {BOOKKEEPER} <{BOOKKEEPER_EMAIL}>"
        m = key.replace("monthly-", "")
        email["subject"] = f"Invoices & bills - {MONTH_LABEL.get(m, m)}"
        opener = _monthly_opener_pool.pop()
        signoff = MONTHLY_SIGNOFFS[RNG.randrange(len(MONTHLY_SIGNOFFS))]
        email["paragraphs"] = [opener] + email["paragraphs"] + [signoff]
    elif key.startswith("stmt-"):
        email["from"] = f"{BANK_NAME} eStatements <estatements@hudsonyardstrust.com>"
        email["to"] = f"{FIRM_NAME} <{HALLORAN_EMAIL}>"
        email["subject"] = f"Your {BANK_NAME} statement is available"
        opener = STMT_OPENERS[RNG.randrange(len(STMT_OPENERS))]
        email["paragraphs"] = [opener] + email["paragraphs"]
    elif key.startswith("expense-log-"):
        email["from"] = f"{VANCE_NAME} <{VANCE_EMAIL}>"
        email["to"] = f"{HALLORAN_NAME} <{HALLORAN_EMAIL}>, {BOOKKEEPER} <{BOOKKEEPER_EMAIL}>"
        email["subject"] = "Expense log + receipts"
        opener = EXPENSE_LOG_OPENERS[RNG.randrange(len(EXPENSE_LOG_OPENERS))]
        email["paragraphs"] = [opener] + email["paragraphs"] + ["-- Owen"]


for key in BATCH_DIR:
    default_fill_batch(key)


def write_body_txt(key: str) -> None:
    email = BATCH_EMAIL[key]
    date_ = BATCH_DATE[key]
    lines = [
        f"From: {email['from']}",
        f"To: {email['to']}",
        f"Date: {date_}",
        f"Subject: {email['subject']}",
        "",
    ]
    lines.extend(email["paragraphs"])
    if email["attachments"]:
        lines.append("")
        lines.append("Attachments:")
        for fname, note in email["attachments"]:
            lines.append(f"- {fname} ({note})")
    text = "\n".join(lines) + "\n"
    out_path = batch_path(key, "body.txt")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(text)


for key in BATCH_DIR:
    write_body_txt(key)

print("Batch emails written.", file=sys.stderr)

# ---------------------------------------------------------------------------
# Opening position (opening_position.json)
# ---------------------------------------------------------------------------

OPENING_POSITION = {
    "period_start": PERIOD_START,
    "period_end": PERIOD_END,
    "as_of": AS_OF_OPEN,
    "cash_by_account": {
        "1000": {"amount_cents": OPEN_CASH, "doc_ids": [DOC_OPEN]},
    },
    "accounts_receivable": [
        {"debtor": debtor, "amount_cents": amt, "doc_ids": [DOC_OPEN]}
        for debtor, amt in OPEN_AR.items()
    ],
    "accounts_payable": [
        {"creditor": creditor, "amount_cents": amt, "doc_ids": [DOC_OPEN]}
        for creditor, amt in OPEN_AP.items()
    ],
    "equity_components": {
        "Partner Capital - Halloran": {"account_code": "3100", "amount_cents": OPEN_CAP_HALLORAN, "doc_ids": [DOC_OPEN]},
        "Partner Capital - Vance": {"account_code": "3110", "amount_cents": OPEN_CAP_VANCE, "doc_ids": [DOC_OPEN]},
    },
    "other_balances": {},
    "depreciation_policy": "N/A -- Halloran & Vance is a service business with no fixed assets or depreciation.",
}

# ---------------------------------------------------------------------------
# Write the four machine-readable files
# ---------------------------------------------------------------------------

L.write_ledger(os.path.join(LAB_DIR, "ledger.jsonl"), LEDGER)
L.write_documents(os.path.join(LAB_DIR, "documents.jsonl"), DOCUMENTS)
L.write_statements(os.path.join(LAB_DIR, "statements.jsonl"), STATEMENTS)
L.write_opening_position(os.path.join(LAB_DIR, "opening_position.json"), OPENING_POSITION)

print(f"Wrote {len(LEDGER)} ledger lines, {len(DOCUMENTS)} documents, "
      f"{len(STATEMENTS)} statements.", file=sys.stderr)

# ---------------------------------------------------------------------------
# Preflight: scan rendered materials for validate.py's forbidden-document
# regex before calling the run done.
# ---------------------------------------------------------------------------

FORBIDDEN_PATTERNS = [
    r"balance[\s_-]?sheet",
    r"profit[\s_-]?(and|&)[\s_-]?loss",
    r"\bp\s*&\s*l\b",
    r"\btrial[\s_-]?balance\b",
    r"general[\s_-]?ledger",
    r"\bgl[\s_-]?export\b",
    r"year[\s_-]?end[\s_-]?summary",
    r"management[\s_-]?accounts",
    r"tax[\s_-]?return",
    r"form\s*1120",
    r"form\s*1065",
    r"schedule\s*k-?1",
]
_FORBIDDEN_RE = re.compile("|".join(FORBIDDEN_PATTERNS), re.IGNORECASE)


def preflight_scan():
    hits = []
    for dirpath, _, filenames in os.walk(MATERIALS_ROOT):
        for fn in filenames:
            full = os.path.join(dirpath, fn)
            if _FORBIDDEN_RE.search(fn):
                hits.append(f"filename: {full}")
            ext = os.path.splitext(fn)[1].lower()
            text = ""
            try:
                if ext in (".txt",):
                    with open(full, "r", encoding="utf-8", errors="ignore") as f:
                        text = f.read()
                elif ext == ".docx":
                    from docx import Document
                    d = Document(full)
                    text = "\n".join(p.text for p in d.paragraphs)
                elif ext == ".xlsx":
                    from openpyxl import load_workbook
                    wb = load_workbook(full, read_only=True, data_only=True)
                    parts = []
                    for ws in wb.worksheets:
                        for row in ws.iter_rows(values_only=True):
                            parts.append(" ".join(str(c) for c in row if c is not None))
                    text = "\n".join(parts)
                elif ext == ".pdf":
                    from pypdf import PdfReader
                    reader = PdfReader(full)
                    text = "\n".join((p.extract_text() or "") for p in reader.pages)
            except Exception as e:
                hits.append(f"unreadable ({e}): {full}")
                continue
            if text and _FORBIDDEN_RE.search(text):
                hits.append(f"content: {full}")
    return hits


_hits = preflight_scan()
if _hits:
    print("PREFLIGHT FORBIDDEN-PATTERN HITS:", file=sys.stderr)
    for h in _hits:
        print(f"  - {h}", file=sys.stderr)
else:
    print("Preflight forbidden-pattern scan: clean.", file=sys.stderr)

# ---------------------------------------------------------------------------
# Summary
# ---------------------------------------------------------------------------

total_files = 0
total_bytes = 0
fmt_counts: dict[str, int] = {}
for dirpath, _, filenames in os.walk(MATERIALS_ROOT):
    for fn in filenames:
        full = os.path.join(dirpath, fn)
        total_files += 1
        total_bytes += os.path.getsize(full)
        ext = os.path.splitext(fn)[1].lstrip(".").lower() or "txt"
        fmt_counts[ext] = fmt_counts.get(ext, 0) + 1

print(f"\nTotal files: {total_files}  Total bytes: {total_bytes} ({total_bytes/1e6:.2f} MB)", file=sys.stderr)
print(f"Format breakdown: {fmt_counts}", file=sys.stderr)

revenue = L.account_balance_cents(LEDGER, "4000", since=PERIOD_START, as_of=PERIOD_END)
totals = L.balance_sheet_totals(LEDGER, as_of=PERIOD_END)
net_profit = totals.income - totals.expense
print(f"FY revenue: {money_str(revenue)}  Net profit: {money_str(net_profit)}  "
      f"Assets: {money_str(totals.assets)}  Closing cash: "
      f"{money_str(L.account_balance_cents(LEDGER, '1000', as_of=PERIOD_END))}", file=sys.stderr)

_margin = net_profit / revenue
print(f"Net margin (pre-partner-comp): {_margin:.1%}", file=sys.stderr)
assert 0.20 <= _margin <= 0.30, f"net margin {_margin:.3f} out of the 20-30% target band"

# Per-counterparty subcontractor spend -- sanity check that no single named
# sub-consultant carries an implausible share of the year.
_sub_by_vendor: dict[str, int] = {}
for l in LEDGER:
    if l["account_code"] == "6040" and PERIOD_START <= l["date"] <= PERIOD_END:
        _sub_by_vendor[l["counterparty"]] = _sub_by_vendor.get(l["counterparty"], 0) + l["debit"]
print("Subcontractor Expense (6040) by vendor, FY:", file=sys.stderr)
for v, amt in sorted(_sub_by_vendor.items(), key=lambda kv: -kv[1]):
    print(f"  {v}: {money_str(amt)}", file=sys.stderr)

# Monthly running cash balance -- confirm the operating account never goes
# negative under the heavier expense load.
_running = OPEN_CASH
print("Monthly closing cash balance (account 1000):", file=sys.stderr)
for m in MONTHS:
    _month_start, _month_end = month_bounds(m)
    _month_lines = [l for l in LEDGER if l["account_code"] == "1000" and _month_start <= l["date"] <= _month_end]
    for l in sorted(_month_lines, key=lambda l: l["date"]):
        _running += l["debit"] - l["credit"]
    flag = "  <-- NEGATIVE" if _running < 0 else ""
    print(f"  {m}: {money_str(_running)}{flag}", file=sys.stderr)
    if _running < 0:
        print(f"WARNING: cash goes negative in {m}", file=sys.stderr)

# ---------------------------------------------------------------------------
# answer-key.md -- figures computed from the ledger, never hand-typed.
# ---------------------------------------------------------------------------

EXPENSE_CODES = ["6000", "6010", "6020", "6030", "6040", "6050", "6060", "6070",
                  "6080", "6130", "6140", "6150", "6900"]

tb_asof = L.trial_balance(LEDGER, as_of=PERIOD_END)
income_amt = tb_asof.get("4000", 0)
expense_by_code = {c: tb_asof.get(c, 0) for c in EXPENSE_CODES}
total_expense = sum(expense_by_code.values())
net_income_final = income_amt - total_expense

cash_bal = tb_asof.get("1000", 0)
ar_bal = tb_asof.get("1200", 0)
ap_bal = tb_asof.get("2000", 0)
accrued_payroll_bal = tb_asof.get("2200", 0)
cap_h = tb_asof.get("3100", 0)
cap_v = tb_asof.get("3110", 0)
draws_h = tb_asof.get("3120", 0)
draws_v = tb_asof.get("3130", 0)
total_assets = cash_bal + ar_bal
total_liabilities = ap_bal + accrued_payroll_bal
# 3120/3130 are contra-equity (debit-normal); with no closing entries they
# stand open at period end and must be netted against capital here exactly
# as lib/ledger.py's balance_sheet_totals nets them by account type.
total_equity_pre_income = cap_h + cap_v - draws_h - draws_v
total_equity_with_income = total_equity_pre_income + net_income_final

halloran_share = round(net_income_final * 0.60)
vance_share = net_income_final - halloran_share

assert total_assets == total_liabilities + total_equity_with_income, "balance sheet does not balance"

ar_by_cp: dict[str, int] = {}
for l in LEDGER:
    if l["account_code"] == "1200" and l["date"] <= PERIOD_END:
        ar_by_cp[l["counterparty"]] = ar_by_cp.get(l["counterparty"], 0) + l["debit"] - l["credit"]
ar_unpaid = {k: v for k, v in ar_by_cp.items() if v > 0}

ap_by_cp: dict[str, int] = {}
for l in LEDGER:
    if l["account_code"] == "2000" and l["date"] <= PERIOD_END:
        ap_by_cp[l["counterparty"]] = ap_by_cp.get(l["counterparty"], 0) + l["credit"] - l["debit"]
ap_unpaid = {k: v for k, v in ap_by_cp.items() if v > 0}

# Real examples of the three date-format variants used for issued_date,
# pulled straight from the invoice data rather than re-typed.
_date_examples = []
for idx, (mkey, ckey, phase, amt_dollars, day, paid_date, as_docx) in enumerate(INVOICES[:3]):
    inv_num = f"HV-{INV_NUM_START + idx}"
    issue_iso = f"{mkey}-{day:02d}"
    _date_examples.append((inv_num, fmt_date_variant(issue_iso, idx % 3)))

pmt_dup_pdf = f"{MATERIALS_REL}/{BATCH_DIR['monthly-2024-09']}/tribeca_paper_receipt.pdf"
pmt_dup_jpg = f"{MATERIALS_REL}/{BATCH_DIR['expense-log-1']}/receipt_tribeca_paper_print_co_2024-09-11.jpg"
pmt_personal = f"{MATERIALS_REL}/{BATCH_DIR['monthly-2025-02']}/atlanticcrest_receipt_personal.pdf"
pmt_handwritten = f"{MATERIALS_REL}/{BATCH_DIR['monthly-2025-04']}/chelsea_hardware_receipt.jpg"
pmt_capital = f"{MATERIALS_REL}/{BATCH_DIR['monthly-2024-12']}/vance_capital_contribution_note.txt"
pmt_opening = f"{MATERIALS_REL}/{BATCH_DIR['opening']}/opening_letter.pdf"

ar_lines_md = "\n".join(f"- {k}: {money_str(v)}" for k, v in sorted(ar_unpaid.items()))
ap_lines_md = "\n".join(f"- {k}: {money_str(v)}" for k, v in sorted(ap_unpaid.items()))

# Individual unpaid bills at period end (bill-count basis, not just
# vendor-count), computed from BILLS rather than re-typed.
_unpaid_bills = [(vendor, issue_date, amt_dollars) for (vendor, issue_date, amt_dollars, paid_date, code) in BILLS
                  if paid_date is None and issue_date <= PERIOD_END]
_unpaid_bills.sort(key=lambda b: b[1])
unpaid_bill_count = len(_unpaid_bills)
unpaid_bills_md = "\n".join(f"- {v} (issued {d}): {money_str(cents(a))}" for v, d, a in _unpaid_bills)

trial_balance_rows = []
for code in sorted(tb_asof):
    meta = L.CHART[code]
    bal = tb_asof[code]
    if meta["normal_side"] == "debit":
        debit_col, credit_col = (bal, 0) if bal >= 0 else (0, -bal)
    else:
        debit_col, credit_col = (0, bal) if bal >= 0 else (-bal, 0)
    if debit_col == 0 and credit_col == 0:
        continue
    trial_balance_rows.append((code, meta["name"], debit_col, credit_col))
tb_debit_total = sum(r[2] for r in trial_balance_rows)
tb_credit_total = sum(r[3] for r in trial_balance_rows)

tb_table_md = "\n".join(
    f"| {code} | {name} | {money_str(d) if d else ''} | {money_str(c) if c else ''} |"
    for code, name, d, c in trial_balance_rows
)

pnl_table_md = "\n".join(
    f"| {L.CHART[code]['name']} | {money_str(amt)} |" for code, amt in expense_by_code.items() if amt
)

ANSWER_KEY = f"""# Halloran & Vance Design Partners -- Answer Key (Challenge Three)

Fiscal year ended 30 June 2025 (period {PERIOD_START} to {PERIOD_END}). Source
documents span July 2024 through August 2025 (fourteen months); the
July-August 2025 trailing stub is bank-reconciled and document-evidenced
like every other month but is **excluded** from every figure in this answer
key and from the period balance sheet / P&L below (validator check 7 scopes
it out the same way).

## 1. Profit-sharing ratio and the trailing stub

Per the prior CPA firm's closing letter (`{pmt_opening}`), profits and
losses are shared **60% to Margaret Halloran / 40% to Owen Vance**. This
ratio is informational prose from the opening letter; it is not journaled
as an automatic allocation of current-year income into the partners'
capital accounts anywhere in `ledger.jsonl` (doing so would require a
current-period-derived net-income figure to be committed to a ledger
entry, which is fine internally, but the point of the ratio for this
exercise is descriptive). If a reader wants to state each partner's implied
share of FY2025 net income, it is illustrative, not booked:
Halloran {money_str(halloran_share)} (60%), Vance {money_str(vance_share)} (40%).

The trailing stub (July-August 2025) contains two invoices, one payroll pay
run, ordinary monthly rent/utilities/telephone/insurance/bank-fee/storage
entries, a handful of subcontractor bills, and both partners' draws for
those two months, all bank-reconciled in the July and August 2025
statements. None of it is included in the FY2025 revenue, expense, P&L,
balance sheet, or trial balance figures below -- validator check 7
evaluates strictly as of `period_end` (2025-06-30), and the stub's own
draws/expenses are dated after that, so they never enter the FY2025
accounting identity even though 3120/3130 keep accumulating through August.

## 2. Opening AR/AP settlement treatment

The prior CPA's closing letter as at 30 June 2024 states three opening
receivables (Bellcourt Retail Group, Ansel Family Residence, Larkspur
Hospitality LLC; combined {money_str(sum(OPEN_AR.values()))}) and three
opening payables (Reyes Drafting Studio, Ionescu Lighting Consultants,
Tribeca Paper & Print Co; combined {money_str(sum(OPEN_AP.values()))}).
Each is collected/paid in full within the first quarter of the period
(entries `OBSETTLE-AR-*` / `OBSETTLE-AP-*` in `ledger.jsonl`, dated
July-September 2024) and appears as an ordinary deposit/withdrawal on the
relevant month's bank statement. None of these six settlements touch
Sales Revenue (4000) or any expense account -- they clear a receivable or
payable that was already recognized as at 30 June 2024, so booking them
to income/expense in FY2025 would double count revenue/expense already
implicitly recognized in the prior period. Evidenced by the opening letter
(`{pmt_opening}`) plus the relevant month's bank statement.

## 3. Mandated defects

### Defect 1 -- duplicate receipt (same purchase, two formats)

The {money_str(dup_amount)} {OFFICE_SUPPLY_VENDOR} purchase on
{dup_date} is shipped **twice**: once as a clean text-native PDF receipt
emailed the same month (`{pmt_dup_pdf}`), and again a month later as a
photographed JPG attached to the September self-maintained expense log
(`{pmt_dup_jpg}`). There is exactly **one** real ledger entry for this
expense (`SUPPLIES-001`, debit 6050 / credit 1000, {dup_date}), which cites
the PDF receipt's `doc_id` plus that month's bank statement. The
photographed JPG is a real shipped file but is deliberately **not**
referenced by any ledger line --
a naive ingestion pass that treats every shipped receipt as a distinct
transaction will double-count this {money_str(dup_amount)}. Correct
treatment: one expense, {money_str(dup_amount)}, Office Supplies Expense.

### Defect 2 -- personal expense on the business account

A {money_str(personal_amount)} Atlantic Crest Airlines round-trip (Margaret
Halloran's personal weekend trip to Miami, {personal_date}) was charged to
the firm's operating account. Receipt at `{pmt_personal}`; the same
month's email batch (`{MATERIALS_REL}/{BATCH_DIR['monthly-2025-02']}/body.txt`) has
Margaret flagging it explicitly. Correct treatment: **not** a business
expense -- booked as a debit to Partner Draws - Halloran (3120), credit
Cash (1000), entry `PERSONAL-001`. It stands in 3120 alongside her other
FY2025 monthly draws (no closing entry moves it anywhere); see &sect;4 for
how the open draws balance is netted into equity at period end.

### Defect 4 -- AR and AP unpaid at period end

At least four sales invoices and at least four vendor bills remain unpaid
as of 30 June 2025.

Unpaid sales invoices (accounts receivable, {len(ar_unpaid)} clients,
{money_str(sum(ar_unpaid.values()))} total):
{ar_lines_md}

Unpaid vendor bills (accounts payable, by vendor; {len(ap_unpaid)} vendors
carrying open balances):
{ap_lines_md}

Individually, {unpaid_bill_count} separate bills are unpaid at period end
(most of them pages inside a monthly bundled attachment rather than their
own file -- see the `vendor_bills_<month>.pdf` documents):
{unpaid_bills_md}

This satisfies the "at least four vendor bills unpaid" requirement on a
bill-count basis as well as a vendor-count basis. Correct treatment: these
remain live Accounts Receivable (1200) / Accounts Payable (2000) balances
at period end; revenue and expense were already recognized on invoice/bill
issuance, so no further P&L impact at collection/payment.

### Defect 6 -- inconsistent date formats across sources

`documents.jsonl`'s `issued_date` field (and the documents themselves)
deliberately vary date format by rotating through three styles as invoices
and bills are issued, e.g. invoice {_date_examples[0][0]} shows
`{_date_examples[0][1]}`, invoice {_date_examples[1][0]} shows
`{_date_examples[1][1]}`, and invoice {_date_examples[2][0]} shows
`{_date_examples[2][1]}` -- MM/DD/YYYY, ISO YYYY-MM-DD, and "D Mon YYYY"
all appear across the corpus (bills-in rotate through the same three
styles independently). Every date inside `ledger.jsonl`, `statements.jsonl`,
and `opening_position.json` is, as required, always ISO with no
exceptions; only the rendered documents' own printed dates and
`documents.jsonl.issued_date` vary.

### Defect 7 -- handwritten-looking cash receipt, photographed at an angle

A {money_str(hw_amount)} hardware purchase from {HARDWARE_VENDOR} on
{hw_date} is evidenced by a handwritten-style receipt image, rendered via
`handwritten_note_image()` and passed through the photograph pass (slight
rotation, perspective warp, uneven lighting): `{pmt_handwritten}`. Booked
to Repairs & Maintenance Expense (6140), entry `REPAIRS-001`.

### Defects 3, 5, 8, 9, 10 -- not applicable to Halloran & Vance

Inter-account transfers (3) and the cancelled-invoice credit note (5) are
Ferrone Provisions LLC's defects; the credit-card statement (8) and loan
amortization schedule (9) are Bright Harbor Fabrication's. Halloran & Vance
uses a single cash account, has no credit card or loan in its chart of
accounts usage, and issues no credit notes -- none of these apply here.

## 4. Additional structural notes

- **Vendor bill bundling**: most months' subcontractor and vendor bills do
  not arrive as one file per bill -- they are scanned/rendered together
  into a single `vendor_bills_<month>.pdf` (kind `multi_document_bundle`),
  the way a real client actually forwards a stack of paperwork in one
  attachment. Each individual bill inside is still its own `ledger.jsonl`
  entry (its own amount, date, and account code); all of that month's bill
  entries simply cite the same bundle `doc_id`. A vendor's bill is page N
  of that file, not its own filename -- check the attachment note in each
  `body.txt` for which vendors are inside before assuming a bundle is a
  single transaction. The November 2024 bundle
  (`{MATERIALS_REL}/{BATCH_DIR['monthly-2024-11']}/vendor_bills_2024-11.pdf`)
  is image-only (scanned) and needs OCR; the rest are text-native PDFs.
- **"Two regular subcontractors" (SPEC &sect;2)**: Reyes Drafting Studio
  (drafting) and Ionescu Lighting Consultants (lighting) are the two
  regular, near-monthly subcontractors. Cobalt Structural Engineering
  PLLC, Rendercraft Visualization Studio, Third Rail Code & Egress
  Consulting, and Bellweather Procurement & FF&E Sourcing are occasional,
  project-linked sub-consultants billing only when a specific phase (most
  often the Whitfield hospitality project) needs their scope -- together
  with Reyes and Ionescu they make up the full Subcontractor Expense
  (6040) balance below.
- **Revenue recognition**: every dollar of Sales Revenue (4000) traces to
  an issued invoice document (`invoice_out`); there is no unbilled/WIP
  revenue anywhere in `ledger.jsonl`.
- **Partner capital contribution**: Owen Vance contributed an additional
  {money_str(capital_amount)} on {capital_date}, evidenced by a deposit on
  that month's bank statement plus a short note (`{pmt_capital}`); entry
  `CAPCONTRIB-001` (debit 1000 / credit 3110).
- **Partner draws**: recurring monthly withdrawals, different per partner
  (Halloran {money_str(DRAW_HALLORAN)}/month, Vance {money_str(DRAW_VANCE)}/month),
  posted to the contra-equity Draws accounts (3120/3130) and left standing
  open at period end -- Halloran {money_str(draws_h)}, Vance
  {money_str(draws_v)} through 2025-06-30. No closing entry moves them; each
  individual draw is evidenced by its own dated bank withdrawal.
  `lib/ledger.py`'s `balance_sheet_totals` (and the trial-balance/equity
  figures below) net every equity-type account by its own debit/credit
  activity, so these debit balances correctly reduce total equity rather
  than inflate it -- no derived closing entry is needed for the accounting
  identity (validator check 7) to hold.
- **Accrued payroll**: {EMPLOYEE_NAME}'s June 2025 salary and employer
  payroll tax (combined {money_str(TOTAL_DRAFT)}) is earned by period end
  but not paid until 2025-07-05 (the payroll provider's regular pay-date
  lag). Accrued at period end to Accrued Payroll Liabilities (2200),
  evidenced by the payroll provider's advance notice
  (`{MATERIALS_REL}/{BATCH_DIR['monthly-2025-06']}/payroll_accrual_notice_june.pdf`), and
  settled in July against that liability, not against expense again.
- **NY sales tax**: not applicable. Halloran & Vance is a professional
  design-services partnership; its fees are not subject to New York sales
  tax, so account 2100 (Sales Tax Payable) does not appear anywhere in
  this company's ledger.

## 5. Profit & Loss -- fiscal year ended 30 June 2025

| Revenue | Amount |
|---|---|
| Sales Revenue | {money_str(income_amt)} |

| Expense | Amount |
|---|---|
{pnl_table_md}
| **Total Expense** | **{money_str(total_expense)}** |

**Net Income: {money_str(net_income_final)}**

## 6. Balance Sheet -- as at 30 June 2025

| Assets | Amount |
|---|---|
| Cash - Operating (1000) | {money_str(cash_bal)} |
| Accounts Receivable (1200) | {money_str(ar_bal)} |
| **Total Assets** | **{money_str(total_assets)}** |

| Liabilities | Amount |
|---|---|
| Accounts Payable (2000) | {money_str(ap_bal)} |
| Accrued Payroll Liabilities (2200) | {money_str(accrued_payroll_bal)} |
| **Total Liabilities** | **{money_str(total_liabilities)}** |

| Equity | Amount |
|---|---|
| Partner Capital - Halloran (3100) | {money_str(cap_h)} |
| Partner Capital - Vance (3110) | {money_str(cap_v)} |
| Partner Draws - Halloran (3120, contra-equity) | ({money_str(draws_h)}) |
| Partner Draws - Vance (3130, contra-equity) | ({money_str(draws_v)}) |
| Current-year net income (undistributed; see &sect;1 for the 60/40 illustrative split) | {money_str(net_income_final)} |
| **Total Equity** | **{money_str(total_equity_with_income)}** |

**Total Liabilities + Equity: {money_str(total_liabilities + total_equity_with_income)}**
(ties to Total Assets above, per validator check 7's accounting identity.)

## 7. Trial Balance -- as at 30 June 2025

| Code | Account | Debit | Credit |
|---|---|---|---|
{tb_table_md}
| | **Totals** | **{money_str(tb_debit_total)}** | **{money_str(tb_credit_total)}** |

## 8. Corpus stats

- Files: {total_files}
- Total bytes: {total_bytes} ({total_bytes/1e6:.2f} MB)
- Format breakdown: {fmt_counts}
- Batches: {len(BATCH_DIR)}
"""

with open(os.path.join(LAB_DIR, "answer-key.md"), "w", encoding="utf-8") as f:
    f.write(ANSWER_KEY)

print("answer-key.md written.", file=sys.stderr)

shutil.rmtree(TMP_DIR, ignore_errors=True)
print("Done.", file=sys.stderr)
