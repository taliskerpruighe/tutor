#!/usr/bin/env python3
"""
validate.py -- FROZEN after Stage A. Do not edit from Stage B.

Usage:
    python lab/challenge-three/validate.py <slug>

Reads, per company:
    lab/challenge-three/<slug>/ledger.jsonl
    lab/challenge-three/<slug>/documents.jsonl
    lab/challenge-three/<slug>/statements.jsonl
    lab/challenge-three/<slug>/opening_position.json

and the shipped corpus at:
    content/21-challenges/materials/challenge-three/<slug>/

and runs the nine checks from SPEC.md, printing each by NAME with
PASS/FAIL and the offending items. Exits non-zero if any check fails.

Environment overrides (used by fixtures/run.sh to demonstrate checks
firing without touching the real repo tree):
    CHALLENGE3_LAB_DIR          default: <repo_root>/lab/challenge-three/<slug>
    CHALLENGE3_MATERIALS_ROOT   default: <repo_root>/content/21-challenges/materials/challenge-three/<slug>
    CHALLENGE3_REPO_ROOT        default: repo root (parent of lab/challenge-three/..)
    CHALLENGE3_BASELINE_MANIFEST default: /tmp/spike006-pre.txt
    CHALLENGE3_ALLOWED_TREES    default: content/21-challenges/materials/challenge-three,lab/challenge-three
                                 (comma-separated, relative to repo root)
"""

from __future__ import annotations

import datetime as dt
import json
import os
import re
import sys

THIS_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, THIS_DIR)

from lib import ledger as L  # noqa: E402


# ---------------------------------------------------------------------------
# Path resolution
# ---------------------------------------------------------------------------

def default_repo_root() -> str:
    # lab/challenge-three/validate.py -> repo root is two levels up
    return os.path.abspath(os.path.join(THIS_DIR, "..", ".."))


def resolve_paths(slug: str) -> dict:
    repo_root = os.environ.get("CHALLENGE3_REPO_ROOT", default_repo_root())
    lab_dir = os.environ.get("CHALLENGE3_LAB_DIR", os.path.join(THIS_DIR, slug))
    materials_root = os.environ.get(
        "CHALLENGE3_MATERIALS_ROOT",
        os.path.join(repo_root, "content", "21-challenges", "materials", "challenge-three", slug),
    )
    baseline_manifest = os.environ.get("CHALLENGE3_BASELINE_MANIFEST", "/tmp/spike006-pre.txt")
    allowed_trees_raw = os.environ.get(
        "CHALLENGE3_ALLOWED_TREES",
        "content/21-challenges/materials/challenge-three,lab/challenge-three",
    )
    allowed_trees = [t.strip() for t in allowed_trees_raw.split(",") if t.strip()]
    return {
        "repo_root": repo_root,
        "lab_dir": lab_dir,
        "materials_root": materials_root,
        "baseline_manifest": baseline_manifest,
        "allowed_trees": allowed_trees,
    }


# ---------------------------------------------------------------------------
# Loading
# ---------------------------------------------------------------------------

def load_context(slug: str) -> dict:
    paths = resolve_paths(slug)
    lab_dir = paths["lab_dir"]

    ledger_lines = L.read_ledger(os.path.join(lab_dir, "ledger.jsonl"))
    documents = L.read_documents(os.path.join(lab_dir, "documents.jsonl"))
    statements = L.read_statements(os.path.join(lab_dir, "statements.jsonl"))
    opening_position = L.read_opening_position(os.path.join(lab_dir, "opening_position.json"))

    documents_by_id = {d["doc_id"]: d for d in documents}

    period_start = opening_position["period_start"]
    period_end = opening_position["period_end"]

    return {
        **paths,
        "slug": slug,
        "ledger": ledger_lines,
        "documents": documents,
        "documents_by_id": documents_by_id,
        "statements": statements,
        "opening_position": opening_position,
        "period_start": period_start,
        "period_end": period_end,
    }


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def doc_exists_on_disk(ctx: dict, doc: dict) -> bool:
    p = os.path.join(ctx["repo_root"], doc["path"])
    return os.path.isfile(p)


def line_evidenced(ctx: dict, line: dict) -> bool:
    doc_ids = line.get("doc_ids") or []
    if not doc_ids:
        return False
    for did in doc_ids:
        doc = ctx["documents_by_id"].get(did)
        if doc is None:
            return False
        if not doc_exists_on_disk(ctx, doc):
            return False
    return True


def is_cash_account(code: str) -> bool:
    return code in L.CASH_ACCOUNTS


def statement_account_codes(ctx: dict) -> set[str]:
    return {s["account_code"] for s in ctx["statements"]}


def date_plus_one(iso_date: str) -> str:
    d = dt.date.fromisoformat(iso_date)
    return (d + dt.timedelta(days=1)).isoformat()


# ---------------------------------------------------------------------------
# Check 1: bank statement lines <-> ledger entries, both directions
# ---------------------------------------------------------------------------

def check_1(ctx: dict) -> tuple[bool, list[str]]:
    problems = []
    cash_codes = statement_account_codes(ctx)
    period_start = ctx["period_start"]

    # index statement lines by (account_code, entry_id)
    stmt_index: dict[tuple[str, str], list[dict]] = {}
    for s in ctx["statements"]:
        for ln in s["lines"]:
            key = (s["account_code"], ln["entry_id"])
            stmt_index.setdefault(key, []).append(ln)

    # index ledger cash lines by (account_code, entry_id)
    ledger_index: dict[tuple[str, str], list[dict]] = {}
    for l in ctx["ledger"]:
        if l["account_code"] in cash_codes:
            ledger_index.setdefault((l["account_code"], l["entry_id"]), []).append(l)

    # direction 1: ledger cash line -> must have a statement line
    # (OB opening-balance entries, dated before period_start, precede any
    # statement and are exempt -- they are evidenced via the opening letter
    # instead; see SPEC.md.)
    for (acct, eid), lines in ledger_index.items():
        for l in lines:
            if l["date"] < period_start:
                continue
            matches = stmt_index.get((acct, eid), [])
            cents = l.get("debit", 0) or l.get("credit", 0)
            if not any(m["amount"] == cents for m in matches):
                problems.append(
                    f"ledger entry {eid} on account {acct} ({l['date']}, {cents}c) "
                    f"has no matching statement line"
                )

    # direction 2: statement line -> must have a ledger cash line
    for (acct, eid), lines in stmt_index.items():
        ledger_matches = ledger_index.get((acct, eid), [])
        for sl in lines:
            if not any(lm.get("debit", 0) == sl["amount"] or lm.get("credit", 0) == sl["amount"]
                       for lm in ledger_matches):
                problems.append(
                    f"statement line entry_id={eid} account={acct} "
                    f"({sl['date']}, {sl['amount']}c) has no matching ledger entry"
                )

    return (len(problems) == 0, problems)


# ---------------------------------------------------------------------------
# Check 2: statement arithmetic
# ---------------------------------------------------------------------------

def check_2(ctx: dict) -> tuple[bool, list[str]]:
    problems = []
    for s in ctx["statements"]:
        credits = sum(ln["amount"] for ln in s["lines"] if ln["direction"] == "in")
        debits = sum(ln["amount"] for ln in s["lines"] if ln["direction"] == "out")
        computed = s["opening_balance"] + credits - debits
        if computed != s["closing_balance"]:
            problems.append(
                f"statement {s['stmt_id']} ({s['account_code']}, "
                f"{s['stmt_period_start']}..{s['stmt_period_end']}): "
                f"opening {s['opening_balance']} + in {credits} - out {debits} "
                f"= {computed}, stated closing {s['closing_balance']}"
            )
    return (len(problems) == 0, problems)


# ---------------------------------------------------------------------------
# Check 3: closing[N] == opening[N+1], per account, no gaps
# ---------------------------------------------------------------------------

def check_3(ctx: dict) -> tuple[bool, list[str]]:
    problems = []
    by_account: dict[str, list[dict]] = {}
    for s in ctx["statements"]:
        by_account.setdefault(s["account_code"], []).append(s)

    for acct, stmts in by_account.items():
        stmts.sort(key=lambda s: s["stmt_period_start"])
        for i in range(len(stmts) - 1):
            cur, nxt = stmts[i], stmts[i + 1]
            if cur["closing_balance"] != nxt["opening_balance"]:
                problems.append(
                    f"{acct}: closing of {cur['stmt_id']} ({cur['closing_balance']}) "
                    f"!= opening of {nxt['stmt_id']} ({nxt['opening_balance']})"
                )
            expected_next_start = date_plus_one(cur["stmt_period_end"])
            if nxt["stmt_period_start"] != expected_next_start:
                problems.append(
                    f"{acct}: gap between {cur['stmt_id']} (ends {cur['stmt_period_end']}) "
                    f"and {nxt['stmt_id']} (starts {nxt['stmt_period_start']}), "
                    f"expected start {expected_next_start}"
                )
    return (len(problems) == 0, problems)


# ---------------------------------------------------------------------------
# Check 4: inter-account transfers recognised on both sides, income/expense
# to neither
# ---------------------------------------------------------------------------

def check_4(ctx: dict) -> tuple[bool, list[str]]:
    problems = []
    cash_codes = statement_account_codes(ctx)
    if len(cash_codes) < 2:
        return (True, ["N/A: fewer than two statement-bearing cash accounts; no inter-account transfers possible"])

    period_start = ctx["period_start"]
    by_entry: dict[str, list[dict]] = {}
    for l in ctx["ledger"]:
        by_entry.setdefault(l["entry_id"], []).append(l)

    stmt_entry_ids = {
        (s["account_code"], ln["entry_id"])
        for s in ctx["statements"] for ln in s["lines"]
    }

    for eid, lines in by_entry.items():
        # Opening-balance entries (dated before period_start) legitimately
        # set up multiple cash accounts at once; they are not transfers and
        # precede any statement -- see the same exemption in check 1.
        if all(l["date"] < period_start for l in lines):
            continue
        touched_cash = {l["account_code"] for l in lines if l["account_code"] in cash_codes}
        if len(touched_cash) < 2:
            continue
        # This entry moves money between >=2 of this company's own cash
        # accounts -- it must be a pure transfer: every line on a cash
        # account, nothing on income/expense.
        for l in lines:
            meta = L.account_meta(l["account_code"])
            if meta["type"] in ("income", "expense"):
                problems.append(
                    f"entry {eid} touches cash accounts {sorted(touched_cash)} "
                    f"AND posts to {meta['type']} account {l['account_code']} "
                    f"({meta['name']}) -- transfer must not touch income/expense"
                )
        # both sides must be traceable to a statement line (check 1 already
        # verifies this generally; re-assert narrowly for the transfer legs)
        for acct in touched_cash:
            if (acct, eid) not in stmt_entry_ids:
                problems.append(
                    f"transfer entry {eid}: leg on account {acct} not found in that "
                    f"account's statement lines"
                )

    return (len(problems) == 0, problems)


# ---------------------------------------------------------------------------
# Check 5: every non-cash entry evidenced by a specific shipped document
# ---------------------------------------------------------------------------

def check_5(ctx: dict) -> tuple[bool, list[str]]:
    problems = []
    cash_codes = statement_account_codes(ctx)
    by_entry: dict[str, list[dict]] = {}
    for l in ctx["ledger"]:
        by_entry.setdefault(l["entry_id"], []).append(l)

    for eid, lines in by_entry.items():
        touches_cash = any(l["account_code"] in cash_codes for l in lines)
        if touches_cash:
            continue  # cash-touching entries are covered by checks 1-4
        for l in lines:
            if not line_evidenced(ctx, l):
                problems.append(
                    f"non-cash entry {eid} line on {l['account_code']} "
                    f"({l['date']}, memo={l.get('memo','')!r}) is not evidenced"
                )
    return (len(problems) == 0, problems)


# ---------------------------------------------------------------------------
# Check 6: explicit, explicitly-empty list of unevidenced ledger entries
# ---------------------------------------------------------------------------

def check_6(ctx: dict) -> tuple[bool, list[str]]:
    unevidenced = []
    for l in ctx["ledger"]:
        if not line_evidenced(ctx, l):
            unevidenced.append(f"{l['entry_id']}/{l['account_code']} ({l['date']})")
    return (len(unevidenced) == 0, unevidenced)


# ---------------------------------------------------------------------------
# Check 7: within period, A = L + E (+ Income - Expense); P&L ties to equity
# movement
# ---------------------------------------------------------------------------

def check_7(ctx: dict) -> tuple[bool, list[str]]:
    problems = []
    unbalanced = L.unbalanced_entries(ctx["ledger"])
    for eid in unbalanced:
        d, c = L.entry_balances(ctx["ledger"])[eid]
        problems.append(f"entry {eid} is not balanced: debits {d} != credits {c}")

    period_end = ctx["period_end"]
    totals = L.balance_sheet_totals(ctx["ledger"], as_of=period_end)
    lhs = totals.assets
    rhs = totals.liabilities + totals.equity + totals.income - totals.expense
    if lhs != rhs:
        problems.append(
            f"as of {period_end}: assets ({lhs}) != liabilities + equity + income - expense "
            f"({totals.liabilities} + {totals.equity} + {totals.income} - {totals.expense} = {rhs})"
        )

    return (len(problems) == 0, problems)


# ---------------------------------------------------------------------------
# Check 8: no forbidden document under materials/, no file outside the two
# permitted trees
# ---------------------------------------------------------------------------

FORBIDDEN_PATTERNS = [
    r"balance[\s_-]?sheet",
    r"profit[\s_-]?(and|&)[\s_-]?loss",
    r"\bp\s*&\s*l\b",
    r"\btrial[\s_-]?balance\b",
    r"general[\s_-]?ledger",
    r"\bgl[\s_-]?export\b",
    r"year[\s_-]?end[\s_-]?summary",
    r"management[\s_-]?accounts",
    r"tax[\s_-]?return",
    r"form\s*1120",
    r"form\s*1065",
    r"schedule\s*k-?1",
]
_FORBIDDEN_RE = re.compile("|".join(FORBIDDEN_PATTERNS), re.IGNORECASE)


# PDF content scan covers the first 10 pages of each PDF (Bright Harbor's
# multi-document bundles run 3-6 unrelated bills per file; 3 pages was not
# enough headroom). This is a stated boundary, not silent partial coverage:
# a forbidden phrase past page 10 of a single PDF would not be caught by
# this check. Raster-only (scanned, non-OCR'd) pages are also not read here
# -- this is a cheap text-layer scan, not an OCR pass.
_PDF_CONTENT_SCAN_PAGES = 10


def _extract_text_cheaply(path: str) -> tuple[str, str | None]:
    """Returns (text, error). `error` is None on success; on failure `text`
    is '' and `error` describes why -- an unreadable file must be reported,
    never silently treated as clean."""
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext == ".txt" or ext == ".csv":
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                return f.read(), None
        if ext == ".pdf":
            from pypdf import PdfReader
            reader = PdfReader(path)
            text = "\n".join((p.extract_text() or "") for p in reader.pages[:_PDF_CONTENT_SCAN_PAGES])
            return text, None
        if ext == ".docx":
            from docx import Document
            doc = Document(path)
            return "\n".join(p.text for p in doc.paragraphs[:200]), None
        if ext == ".xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            out = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(max_row=20, values_only=True):
                    out.append(" ".join(str(c) for c in row if c is not None))
            return "\n".join(out), None
    except Exception as e:
        return "", f"{type(e).__name__}: {e}"
    # Formats with no cheap text layer (jpg/jpeg) are intentionally not
    # scanned for content here -- filename patterns and human review cover
    # them; this is not a failure to report.
    return "", None


def _read_manifest(path: str) -> dict[str, int]:
    out = {}
    if not os.path.isfile(path):
        return out
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.rstrip("\n")
            if not line or "\t" not in line:
                continue
            p, size = line.rsplit("\t", 1)
            try:
                out[p] = int(size)
            except ValueError:
                continue
    return out


def _current_manifest(repo_root: str) -> dict[str, int]:
    out = {}
    for dirpath, dirnames, filenames in os.walk(repo_root):
        dirnames[:] = [d for d in dirnames if d != ".git"]
        for fn in filenames:
            full = os.path.join(dirpath, fn)
            rel = "./" + os.path.relpath(full, repo_root)
            try:
                size = os.path.getsize(full)
            except OSError:
                continue
            out[rel] = size
    return out


def _under_any_tree(rel_path: str, trees: list[str]) -> bool:
    norm = rel_path[2:] if rel_path.startswith("./") else rel_path
    for t in trees:
        t = t.rstrip("/")
        if norm == t or norm.startswith(t + "/"):
            return True
    return False


def check_8(ctx: dict) -> tuple[bool, list[str]]:
    problems = []
    materials_root = ctx["materials_root"]

    if os.path.isdir(materials_root):
        for dirpath, _, filenames in os.walk(materials_root):
            for fn in filenames:
                full = os.path.join(dirpath, fn)
                if _FORBIDDEN_RE.search(fn):
                    problems.append(f"forbidden filename pattern: {full}")
                    continue
                text, error = _extract_text_cheaply(full)
                if error is not None:
                    problems.append(f"unreadable, content not scanned: {full} ({error})")
                    continue
                if text and _FORBIDDEN_RE.search(text):
                    problems.append(f"forbidden content pattern inside: {full}")

    baseline = _read_manifest(ctx["baseline_manifest"])
    if baseline:
        current = _current_manifest(ctx["repo_root"])
        for rel_path in current:
            if rel_path not in baseline and not _under_any_tree(rel_path, ctx["allowed_trees"]):
                problems.append(f"new file outside permitted trees: {rel_path}")
    else:
        problems.append(f"WARNING: baseline manifest not found/empty at {ctx['baseline_manifest']} -- extra-file check skipped")

    return (len(problems) == 0, problems)


# ---------------------------------------------------------------------------
# Check 9: opening letter cash per account == first in-period statement's
# opening balance
# ---------------------------------------------------------------------------

def check_9(ctx: dict) -> tuple[bool, list[str]]:
    problems = []
    cash_by_account = ctx["opening_position"].get("cash_by_account", {})
    by_account: dict[str, list[dict]] = {}
    for s in ctx["statements"]:
        by_account.setdefault(s["account_code"], []).append(s)

    for acct, info in cash_by_account.items():
        stmts = sorted(by_account.get(acct, []), key=lambda s: s["stmt_period_start"])
        if not stmts:
            problems.append(f"account {acct}: no statements found at all")
            continue
        first = stmts[0]
        if info["amount_cents"] != first["opening_balance"]:
            problems.append(
                f"account {acct}: opening letter cash {info['amount_cents']} != "
                f"first statement ({first['stmt_id']}) opening balance {first['opening_balance']}"
            )
    return (len(problems) == 0, problems)


# ---------------------------------------------------------------------------
# Runner
# ---------------------------------------------------------------------------

CHECKS = [
    ("check_1_bank_statement_ledger_bidirectional_trace", check_1),
    ("check_2_statement_arithmetic", check_2),
    ("check_3_month_to_month_continuity", check_3),
    ("check_4_inter_account_transfers", check_4),
    ("check_5_noncash_entries_evidenced", check_5),
    ("check_6_unevidenced_ledger_entries_empty", check_6),
    ("check_7_period_balance_sheet_and_pnl_tie", check_7),
    ("check_8_forbidden_documents_and_extra_files", check_8),
    ("check_9_opening_letter_cash_ties_to_statements", check_9),
]


def main(argv: list[str]) -> int:
    if len(argv) != 2:
        print("usage: python validate.py <slug>", file=sys.stderr)
        return 2
    slug = argv[1]
    ctx = load_context(slug)

    all_pass = True
    for name, fn in CHECKS:
        ok, items = fn(ctx)
        status = "PASS" if ok else "FAIL"
        print(f"[{status}] {name}")
        for item in items:
            print(f"    - {item}")
        if not ok:
            all_pass = False

    print()
    print("RESULT:", "PASS" if all_pass else "FAIL")
    return 0 if all_pass else 1


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
